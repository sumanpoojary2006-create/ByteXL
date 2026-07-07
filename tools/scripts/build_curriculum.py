import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ---------- uniform blue palette ----------
BANNER='FF1F3B5C'; SUBNOTE='FFEAF1F8'; HEADER='FF2E5E8C'; CHIP='FF4F7CB0'
GROUP1='FFB9D0EA'; GROUP2='FFD2E2F2'; HILITE='FFCFE0F2'; ZEBRA='FFEEF4FB'; WHITE='FFFFFFFF'
TEXT='FF1B2A3A'; SUBTEXT='FF51606F'; WHITEF='FFFFFFFF'
FONT='Calibri'
thin=Side(style='thin', color='FFCFD8E3'); BORDER=Border(left=thin,right=thin,top=thin,bottom=thin)

def cell(ws,r,c,val,fill=WHITE,font=TEXT,bold=False,italic=False,size=11,wrap=True,halign='left',valign='top',border=True):
    cc=ws.cell(r,c,val); cc.fill=PatternFill('solid',fgColor=fill)
    cc.font=Font(name=FONT,size=size,bold=bold,italic=italic,color=font)
    cc.alignment=Alignment(horizontal=halign,vertical=valign,wrap_text=wrap)
    if border: cc.border=BORDER
    return cc

def sheet_header(ws,title,subtitle,headers,widths):
    ncol=len(headers)
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=ncol)
    cell(ws,1,1,title,fill=BANNER,font=WHITEF,bold=True,size=15,halign='left',valign='center')
    for c in range(2,ncol+1): cell(ws,1,c,None,fill=BANNER,border=False)
    ws.row_dimensions[1].height=30
    ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=ncol)
    cell(ws,2,1,subtitle,fill=SUBNOTE,font=SUBTEXT,italic=True,halign='left',valign='center')
    for c in range(2,ncol+1): cell(ws,2,c,None,fill=SUBNOTE,border=False)
    ws.row_dimensions[2].height=24
    for i,h in enumerate(headers): cell(ws,4,i+1,h,fill=HEADER,font=WHITEF,bold=True,halign='left',valign='center')
    ws.row_dimensions[4].height=26
    for i,w in enumerate(widths): ws.column_dimensions[openpyxl.utils.get_column_letter(i+1)].width=w
    ws.freeze_panes=ws.cell(5,1)
    return 5

def slug(s):
    s=re.sub(r'[^a-zA-Z0-9 ]','',s).strip().lower()
    return re.sub(r'\s+','-',s)

# =====================================================================================
#  SEMESTER 1 - 13 UNITS
# =====================================================================================
SEM1_UNITS=[
 ("Introduction to Programming","Think like a programmer before writing Python: turn real problems into clear logic.",
  ["What is Programming? Why Programming Matters","How Computers Work: Inputs, Processing, and Outputs (IPO)",
   "Computational Thinking: Decomposition, Pattern Recognition, Abstraction","Problem-Solving Approach: From Problem Statement to Plan",
   "Algorithms: Designing Step-by-Step Logic","Pseudocode: Writing Logic Before Code","Flowcharts: Visualizing Program Flow and Decisions",
   "Why Python? Readability, Ecosystem, and Real-World Uses","Setting Up Python and Your First Program (REPL, scripts, print())"]),
 ("Data Types and Operators","Store and transform information using Python's core value types.",
  ["Variables, Assignment, and Naming Conventions (PEP 8)","Numbers: int, float, and complex","Strings and Booleans (introduction)",
   "Type Conversion (Casting) and type()","Arithmetic Operators","Comparison (Relational) Operators","Logical Operators: and, or, not",
   "Assignment and Augmented Operators","Operator Precedence and Associativity","Input and Output: input(), print(), and f-strings"]),
 ("Control Flow","Make programs decide and respond using branching logic.",
  ["Decision-Making: Why Branching Matters","The if Statement","if-else: Two-Way Decisions","elif: Multi-Way Branching",
   "Nested Conditions and Guard Clauses","match-case: Python's switch-style Structure","Truthiness and Boolean Logic in Conditions",
   "Conditional (Ternary) Expressions","Input Validation with Conditions"]),
 ("Looping","Automate repetition and process collections of data efficiently.",
  ["Why Loops: Repetition and Automation","while Loops: Counters and Sentinels","for Loops and range()","Iterating over Sequences and Strings",
   "break, continue, and the loop-else Clause","Nested Loops: Grids and Patterns","Common Loop Patterns: Count, Sum, Min/Max, Search",
   "Avoiding Infinite Loops and Off-by-One Errors"]),
 ("Strings","Work with text in depth: the most common data type students will touch.",
  ["What Is a String and How It Is Stored","Indexing and Slicing","String Immutability","Common String Methods (upper, lower, strip, replace)",
   "Splitting and Joining (split, join)","Searching Within Strings (in, find, count)","String Formatting and f-strings",
   "Escape Sequences and Multi-line Strings","Practical Text Processing"]),
 ("Lists and Tuples","Order, store, and transform sequences of values.",
  ["Lists: Creation, Indexing, and Slicing","Mutating Lists (append, insert, remove, pop)","List Methods and Sorting","List Comprehensions",
   "Tuples: Immutability and Packing/Unpacking","When to Use a Tuple vs a List","Nested Lists","Iterating over Lists and Tuples"]),
 ("Sets and Dictionaries","Handle uniqueness and key-value data, the workhorses of real programs.",
  ["Sets: Uniqueness and Creation","Set Operations (union, intersection, difference)","Dictionaries: Key-Value Pairs",
   "Accessing, Adding, Updating, and Removing Items","Iterating Dictionaries (keys, values, items)","Dictionary Comprehensions",
   "Nested Dictionaries","Choosing the Right Data Structure"]),
 ("Functions","Package logic into reusable, well-named building blocks.",
  ["Why Functions: Reuse and Decomposition","Defining and Calling Functions","Parameters, Arguments, and Return Values",
   "Default, Keyword, and Positional Arguments","*args and **kwargs","Lambda (Anonymous) Functions",
   "map, filter, and reduce (right after lambda)","Useful Built-in Functions: all, any, len, sum, sorted, min, max",
   "Nested Functions","Variable Scope and the LEGB Rule (follows the variables unit)","Recursion: Base Case and Recursive Step","Docstrings and Clean Function Design"]),
 ("Basic Object-Oriented Programming","A gentle first look at modeling real-world things with classes (kept light).",
  ["Why OOP? Modeling Real-World Things","What is a Class? What is an Object?","Creating a Class and Instantiating Objects",
   "Attributes (Instance Variables)","Methods and the self Parameter","The __init__ Constructor","A Simple Class Walkthrough (Student / BankAccount)",
   "Objects vs. Plain Data: When to Use a Class"]),
 ("Modules and Packages","Organize, reuse, and manage code and dependencies the real-world way.",
  ["Why Modules: Organizing and Reusing Code","import, from-import, and Aliases","Touring the Standard Library (math, random, datetime, os)",
   "Creating Your Own Module","Packages and __init__.py","pip and Installing Third-Party Packages (PyPI)",
   "Virtual Environments: Why Isolation Matters (venv)","The uv Package Manager: Fast Environments and Dependencies",
   "requirements and Reproducible Setups (optional)"]),
 ("File Handling","Make data persist beyond a single run by reading and writing files.",
  ["Why Files: Persisting Data Beyond Runtime","Reading Text Files (open, read, readlines)","Writing and Appending to Files",
   "The with Statement (Context-Managed Files)","Working with File Paths using pathlib","Finding Files with glob (Pattern Matching)",
   "Reading and Writing CSV Files","Reading and Writing JSON","Common File Pitfalls: Encoding and Missing Files"]),
 ("Exception Handling","Write robust programs that anticipate and recover from failure.",
  ["What Are Errors? Syntax vs. Runtime vs. Logic","Reading a Traceback Bottom-Up","try and except","Handling Multiple and Specific Exceptions",
   "else and finally","Raising Exceptions with raise","Creating Custom Exceptions (introduction)","Validating Input Defensively",
   "Errors as Part of Robust Program Design"]),
 ("Debugging","Find and fix bugs systematically: the most practical real-world skill.",
  ["The Debugging Mindset: Errors Are Normal","Print Debugging and Reading Output","Using assert for Sanity Checks",
   "The Python Debugger (pdb / IDE breakpoints)","Stepping, Inspecting Variables, and the Call Stack","Logging Basics: logging vs print",
   "Isolating a Minimal Failing Case","Rubber-Duck Debugging and Asking Good Questions","AI-Assisted Debugging (used responsibly)"]),
]

SEM1_WEEKS=[
 (1,1,"IPO model, computational thinking, algorithms, pseudocode, flowcharts, first program","Map tasks to IPO; write pseudocode + a flowchart for an eligibility checker","Set 1 (10 Q): trace logic, fix pseudocode, first programs","Quiz U1 (10 MCQ)","-",1.5,1.5),
 (2,2,"Variables, numeric/str/bool types, casting, operators, precedence, I/O & f-strings","Build a tip / EMI / unit converter using input, operators, f-strings","Set 2 (10 Q): expressions, casting, precedence","Quiz U2 (10 MCQ)","-",1.5,1.5),
 (3,3,"if / if-else / elif, nested conditions, guard clauses, match-case, validation","Grade calculator + menu router using match-case","Set 3 (10 Q): branching, match-case, validation","Quiz U3 (10 MCQ)","-",1.5,1.5),
 (4,4,"while/for, range, break/continue/else, nested loops, loop patterns","Multiplication grid + number-guessing game with attempt limit","Set 4 (10 Q): loop patterns, accumulation, search","Quiz U4 (10 MCQ)","Mini project: Number-Guessing Game",1.0,2.0),
 (5,5,"Strings in depth: indexing, slicing, methods, split/join, formatting, immutability","Build a text cleaner and word/palindrome tools","Set 5 (10 Q): string methods, slicing, formatting","Quiz U5 (10 MCQ)","-",1.5,1.5),
 (6,6,"Lists, slicing, methods, comprehensions, tuples, unpacking, nested lists","Build a to-do list manager (in memory)","Set 6 (10 Q): list ops, comprehensions, tuples","Quiz U6 (10 MCQ)","Mini project: To-Do List Manager",1.0,2.0),
 (7,7,"Sets, set operations, dictionaries, iteration, dict comprehensions, choosing a structure","Word-frequency counter + simple inventory using a dict","Set 7 (10 Q): sets, dict lookups, comprehensions","Quiz U7 (10 MCQ)","Mini project: Contact Book / Inventory",1.0,2.0),
 (8,8,"Define/call, params, lambda, map/filter/reduce, all/any/len, nested functions, scope, recursion","Build a text-analytics toolkit of functions; include one recursive function","Set 8 (10 Q): functions, lambda, map/filter, scope, recursion","Quiz U8 (10 MCQ)","Mini project: Text Analyzer Toolkit",1.0,2.0),
 (9,9,"Why OOP, classes/objects, attributes, methods, self, __init__ (kept light)","Model a Student / BankAccount class with attributes and methods","Set 9 (10 Q): classes, objects, attributes, methods","Quiz U9 (10 MCQ)","Mini project: Student/Account Manager (basic OOP)",1.5,1.5),
 (10,10,"Imports, stdlib tour, own modules, packages, pip, venv, uv basics","Create a package and set up a venv/uv environment; install and use a library","Set 10 (10 Q): imports, modules/packages, env basics","Quiz U10 (10 MCQ)","Mini project: Build a package (Gold Price Index calculator)",1.0,2.0),
 (11,11,"Read/write text, with-statement, pathlib, glob, CSV, JSON, pitfalls","Build a notes/expense logger that persists to CSV and JSON; find files with glob","Set 11 (10 Q): file read/write, glob, CSV/JSON","Quiz U11 (10 MCQ)","Mini project: File-backed Expense Logger",1.0,2.0),
 (12,12,"Errors vs exceptions, tracebacks, try/except/else/finally, raise, custom errors","Add robust error handling and input validation to the expense logger","Set 12 (10 Q): exception handling, custom errors","Quiz U12 (10 MCQ)","-",1.5,1.5),
 (13,13,"Debugging mindset, print/assert, pdb/breakpoints, logging, minimal repro","Debugging clinic: fix seeded broken programs and add logging","Set 13 (10 Q): debugging and log-reading scenarios","Quiz U13 (10 MCQ)","Mini project: Bug-Fix Clinic",1.0,2.0),
 (14,"Semester Project","Integrate Units 1-13 into one application","Build the Semester 1 project (see Projects sheet)","Project rubric-based","-","Semester Project: CLI Expense/Student Manager (build)",0.5,2.5),
 (15,"Review & Final","Consolidation, gaps, exam prep, demos","Timed practical + viva","Final coding assessment","Final MCQ assessment","Semester Project demo + review",1.0,2.0),
]

# =====================================================================================
#  SEMESTER 2 - 14 UNITS
# =====================================================================================
SEM2_UNITS=[
 ("Recap and Python Internals","Bridge from fundamentals and understand how Python actually runs your code.",
  ["Refresher: functions, collections, files, basic OOP, exceptions","How Python Runs Your Code: the Interpreter","From Source to Bytecode",
   "The __pycache__ Folder and .pyc Files","The Import System and Module Resolution","sys.path and How Imports Are Found",
   "Intermediate Files Frameworks Create (FastAPI / Flask context)","Introduction to the Python Data Model (dunder methods)"]),
 ("Encapsulation and Abstraction","Deepen OOP: protect state and hide complexity behind clean interfaces.",
  ["Recap of Classes and Objects","Encapsulation: Bundling Data and Behavior","Access Control: _protected, __private, name-mangling",
   "Properties and Getters/Setters (@property)","Abstraction: Hiding Complexity","Abstract Base Classes (the abc module)","Designing Clean Interfaces"]),
 ("Inheritance and Polymorphism","Model hierarchies and reuse behavior the way real systems do.",
  ["Inheritance Fundamentals","super() and the Constructor Chain","Method Overriding","Polymorphism in Practice",
   "Multiple Inheritance and the MRO","Composition vs. Inheritance","Dunder Methods (__str__, __repr__, __eq__, __len__)",
   "dataclasses","Class, Static, and Instance Methods"]),
 ("Iterators and Generators","Process data lazily and write memory-efficient pipelines.",
  ["The Iterator Protocol (__iter__, __next__)","Iterables vs. Iterators","Building a Custom Iterator","Generators and the yield Keyword",
   "Generator Expressions","Lazy Evaluation and Memory Efficiency","itertools Essentials","When to Use Generators (streaming large data)"]),
 ("Decorators","Extend and wrap behavior cleanly using Python's functional power.",
  ["First-Class Functions and Closures","Functions as Arguments and Return Values","Writing a Simple Decorator","Decorators with Arguments",
   "functools.wraps and Preserving Metadata","Stacking Multiple Decorators","Class Decorators (introduction)","Real-World Decorators: Timing, Caching, Logging"]),
 ("Context Managers","Manage resources safely with guaranteed setup and cleanup.",
  ["The with Statement Revisited","The Context Manager Protocol (__enter__, __exit__)","Writing a Class-Based Context Manager",
   "contextlib and @contextmanager","Managing Resources Safely (files, connections, locks)","Suppressing Exceptions and Guaranteeing Cleanup"]),
 ("Standard Library","Master the batteries-included toolbox before testing and shipping.",
  ["Why the Standard Library Matters (batteries included)","random: Randomness and Sampling",
   "Cryptographic and Security Modules (hashlib, secrets, hmac)","datetime: Dates, Times, and Durations",
   "collections: Counter, defaultdict, namedtuple","itertools: Building Blocks for Iteration","os, sys, and pathlib Essentials","json and csv Recap"]),
 ("Testing","Build confidence and prevent regressions with automated tests.",
  ["Why Test? Confidence and Regression Safety","assert and Manual Testing","Introduction to pytest","Writing and Organizing Test Functions",
   "Fixtures and Setup/Teardown","Parameterized Tests","Mocking and Patching (unittest.mock)","Measuring Test Coverage (coverage.py)",
   "Test-Driven Development (TDD) Workflow"]),
 ("Code Quality and Pre-commit Hooks","Ship professional, consistent, type-checked code automatically.",
  ["Clean Code and PEP 8 in Practice","Type Hints and Static Typing (mypy)","Linting with ruff / flake8","Auto-Formatting with black",
   "Introduction to Git Hooks","The pre-commit Framework","Configuring and Running pre-commit Hooks","Continuous Integration (CI) Basics"]),
 ("Asynchronous Programming","Handle many I/O-bound tasks concurrently with async/await.",
  ["Concurrency vs. Parallelism (mental model)","Blocking vs. Non-Blocking I/O","async and await","The asyncio Event Loop",
   "Coroutines and Tasks","Running Coroutines Concurrently (gather)","Async Context Managers and Iterators","When Async Helps (and When It Does Not)"]),
 ("Multithreading and Multiprocessing","Use threads and processes correctly, GIL and all.",
  ["Processes vs. Threads","The threading Module","The Global Interpreter Lock (GIL)","Thread Safety, Locks, and Race Conditions",
   "multiprocessing for CPU-bound Work","concurrent.futures (Thread and Process Pool Executors)","Choosing Async vs. Threads vs. Processes"]),
 ("CLI Application Development","Turn scripts into real command-line tools people can run.",
  ["Why CLIs: Real Tools Users Run","Reading Arguments with sys.argv","argparse: Arguments, Options, and Subcommands",
   "Building a CLI with typer","Input Validation and Helpful Help Text","Exit Codes and Error Handling in CLIs",
   "Exposing a CLI as a Command (entry points)","UX Best Practices for Command-Line Tools"]),
 ("Database Interaction","Connect Python to real databases and persist data reliably.",
  ["Relational Database Fundamentals","Connecting to a Database from Python","Cursors and Executing Queries",
   "SELECT, INSERT, UPDATE, DELETE from Python","Parameterized Queries (avoiding SQL injection)","Transactions and Commits",
   "Starting with SQLite, then PostgreSQL","Modern Example: PostgreSQL with pgvector for Embeddings"]),
 ("Packaging, Project Structure, and Distribution","Organize, build, and ship a real, installable Python project.",
  ["Professional Project Layout (the src layout)","pyproject.toml and Build Backends","Managing Dependencies and Environments with uv",
   "Building Wheels and Source Distributions","Versioning and Semantic Versioning","Publishing to PyPI (overview)",
   "Documentation and README Essentials","Bringing It Together: A Shippable, Tested Package"]),
]

SEM2_WEEKS=[
 (1,1,"How Python runs code, bytecode, __pycache__/.pyc, import system, framework intermediate files","Inspect .pyc and __pycache__; trace an import; explain what FastAPI/Flask generate","Set 1 (10 Q): interpreter, bytecode, imports","Quiz U1 (10 MCQ)","-",1.5,1.5),
 (2,2,"Encapsulation, access control, @property, abstraction, abstract base classes","Refactor a class to use properties and an abstract base class","Set 2 (10 Q): encapsulation, properties, ABCs","Quiz U2 (10 MCQ)","-",1.5,1.5),
 (3,3,"Inheritance, super, overriding, polymorphism, MRO, dunder methods, dataclasses","Model a domain (shapes/library) with inheritance and dunder methods","Set 3 (10 Q): inheritance, polymorphism, dunders","Quiz U3 (10 MCQ)","Mini project: OOP domain model",1.0,2.0),
 (4,4,"Iterator protocol, generators, yield, generator expressions, itertools","Build a custom iterator and a generator data pipeline","Set 4 (10 Q): iterators, generators, lazy eval","Quiz U4 (10 MCQ)","-",1.5,1.5),
 (5,5,"Closures, decorators, args, functools.wraps, real-world decorators","Write timing/caching/logging decorators and apply them","Set 5 (10 Q): closures and decorators","Quiz U5 (10 MCQ)","Mini project: Reusable decorator utilities",1.0,2.0),
 (6,6,"with revisited, __enter__/__exit__, contextlib, safe resource management","Write a class-based and an @contextmanager context manager","Set 6 (10 Q): context managers and cleanup","Quiz U6 (10 MCQ)","-",1.5,1.5),
 (7,7,"random, hashlib/secrets/hmac, datetime, collections, itertools, os/sys/pathlib","Solve tasks using four or more stdlib modules including random and hashlib","Set 7 (10 Q): standard library modules","Quiz U7 (10 MCQ)","-",1.5,1.5),
 (8,8,"pytest, fixtures, parameterized tests, mocking, coverage, TDD","Write a pytest suite and reach a coverage target by TDD-ing a feature","Set 8 (10 Q): testing and TDD","Quiz U8 (10 MCQ)","Mini project: Test suite + coverage report",1.0,2.0),
 (9,9,"Clean code, type hints/mypy, ruff, black, git hooks, pre-commit, CI","Add type hints and configure pre-commit (ruff + black + mypy) on a repo","Set 9 (10 Q): typing and code quality","Quiz U9 (10 MCQ)","Mini project: Pre-commit-guarded repo",1.0,2.0),
 (10,10,"Concurrency vs parallelism, async/await, event loop, coroutines, gather","Build an async multi-URL/file fetcher","Set 10 (10 Q): asyncio concepts","Quiz U10 (10 MCQ)","-",1.5,1.5),
 (11,11,"Threads vs processes, GIL, locks/races, multiprocessing, concurrent.futures","Speed up an I/O task with threads and a CPU task with processes","Set 11 (10 Q): threads, GIL, processes","Quiz U11 (10 MCQ)","-",1.5,1.5),
 (12,12,"CLIs, sys.argv, argparse, typer, validation, exit codes, entry points","Build a typer CLI with subcommands (the capstone front-end)","Set 12 (10 Q): CLI design, argparse/typer","Quiz U12 (10 MCQ)","Mini project: typer CLI tool (capstone milestone)",1.0,2.0),
 (13,13,"Relational DB, connect, cursor, queries, parameterized queries, transactions, SQLite->PostgreSQL, pgvector","Build a small app backed by SQLite using parameterized queries","Set 13 (10 Q): SQL from Python, cursors, transactions","Quiz U13 (10 MCQ)","Capstone milestone: add database persistence",1.0,2.0),
 (14,14,"src layout, pyproject.toml, uv, wheels, versioning, PyPI, docs","Package the capstone CLI into an installable project with pyproject and uv","Set 14 (10 Q): packaging and distribution","Quiz U14 (10 MCQ)","Capstone milestone: package and ship",1.0,2.0),
 (15,"Capstone & Final","Demos, code review, final assessment","Capstone demo + viva","Final coding assessment","Final MCQ assessment","Capstone demo day + final review",1.0,2.0),
]

READING_BLOCK=("Reading material per unit must include: concept explanation, real-world examples, case studies, beginner-friendly walkthroughs, "
 "runnable code examples (kept minimal in early units), common mistakes, practice prompts, a summary, and key takeaways.")

def reading_for(t): return f"Reading: {t} - concept explanation + real-world analogy, worked examples, common mistakes, practice prompts, summary and key takeaways."

# =====================================================================================
#  WORKBOOK BUILDER
# =====================================================================================
def build(path, sem_no, sem_name, tagline, units, weeks, learner, prereqs, outcomes,
          unit_model_text, recommended_additions, assumptions, cross_note, sem_project, capstone_text):
    wb=openpyxl.Workbook()
    # Overview
    ws=wb.active; ws.title='Overview'
    r=sheet_header(ws,f"Python Curriculum - Semester {sem_no}: {sem_name}",tagline,["Section","Details"],[34,108])
    rows=[("Curriculum title",f"Python {sem_name} (Semester {sem_no})"),
          ("Format","University-style, 15 weeks; runs as either a 3-credit (45 hours) or 4-credit (60 hours) course."),
          ("Structure model","Unit -> Topic (no chapters), following the attached Course Structure format."),
          ("Number of units",f"{len(units)} units across 15 weeks (within the 15-unit ceiling)."),
          ("Target learner",learner),("Prerequisites",prereqs),("Learning outcomes",outcomes),("Continuity",cross_note),
          ("Assessment per unit","10 MCQs (Beginner/Intermediate/Challenge) + 10 course coding questions + 20 question-bank coding questions (1:2 ratio). See 'Assessment & QB Plan'."),
          ("Projects","Four levels: topic exercises, unit mini-projects, a semester project, and a cross-semester capstone. See 'Projects & Capstone'."),
          ("Reading material",READING_BLOCK),
          ("How to use","Teach unit by unit using 'Curriculum (Unit-Topic)'; pace with 'Week-wise Plan'; build question banks from 'Assessment & QB Plan' + 'Question Taxonomy' + 'Templates'.")]
    for i,(a,b) in enumerate(rows):
        cell(ws,r+i,1,a,fill=GROUP1,bold=True); cell(ws,r+i,2,b,fill=(ZEBRA if i%2==0 else WHITE))
        ws.row_dimensions[r+i].height=max(30,16*(1+len(b)//95))
    # Recommended Structure
    ws=wb.create_sheet('Recommended Structure')
    r=sheet_header(ws,"Recommended Curriculum Structure","Why this unit model, how it fits 15 weeks, and how 45h vs 60h differ.",["Topic","Recommendation / Explanation"],[30,112])
    rs=[("Recommended unit model",unit_model_text),
        ("Why not fewer (4-6 units)","Each unit would span multiple weeks and mix unrelated concepts, hurting pacing and making a focused 30-question bank per unit hard."),
        ("Why not 15 units (1/week)","Over-fragments natural topics and leaves no slack for projects, review, and assessment. 13-14 units is the sweet spot under the 15 cap."),
        ("Fit to 15-week calendar","Content units occupy the early and middle weeks; the final weeks are reserved for the semester project or capstone milestones, review, and finals."),
        ("45-hour version (3 credits)","15 weeks x 3 hours = 45 hours. Each week is roughly 1-1.5 theory + 1.5-2 lab hours, covering all units, mini-projects, the semester project, and unit quizzes plus coding challenges."),
        ("60-hour version (4 credits)","15 weeks x 4 hours = 60 hours. The extra hour each week is added to lab/project time, used for deeper code labs, the 20 question-bank coding problems, an extra mini-project per major unit, optional topics, and a more ambitious project/capstone."),
        ("How projects fit","Topic-level exercises every session; unit mini-projects at the end of major units; a semester project; and the cross-semester capstone built across the final weeks of Semester 2."),
        ("How quizzes & coding fit","A 10-MCQ quiz and a 10-question coding challenge close each unit; the 20-question bank set per unit is authored for assessments and practice using the templates."),
        ("Recommended additions beyond feedback",recommended_additions),
        ("Assumptions made",assumptions)]
    for i,(a,b) in enumerate(rs):
        cell(ws,r+i,1,a,fill=GROUP1,bold=True); cell(ws,r+i,2,b,fill=(ZEBRA if i%2==0 else WHITE))
        ws.row_dimensions[r+i].height=max(30,15*(1+len(b)//100))
    # Curriculum (Unit - Topic)
    ws=wb.create_sheet('Curriculum (Unit-Topic)')
    r=sheet_header(ws,f"Semester {sem_no} Curriculum - Unit and Topic","Clean Unit -> Topic structure (no chapters), in teaching order.",["Unit","Topic"],[40,86])
    cur=r
    for ui,(utitle,usum,topics) in enumerate(units,1):
        start=cur
        for ti,t in enumerate(topics):
            cell(ws,cur,2,t,fill=(ZEBRA if ti%2==0 else WHITE)); ws.row_dimensions[cur].height=20; cur+=1
        ws.merge_cells(start_row=start,start_column=1,end_row=cur-1,end_column=1)
        cell(ws,start,1,f"Unit {ui}: {utitle}\n\n{usum}",fill=GROUP1,bold=True,valign='top')
    # Week-wise Plan
    ws=wb.create_sheet('Week-wise Plan')
    r=sheet_header(ws,f"Semester {sem_no} - 15-Week Plan (45h base, +60h add-on)","Theory/Lab hours are the 45-hour base; the 60-hour version adds +1 lab/project hour each week.",
                   ["Week","Unit","Topics Covered","Reading Material","Exercise","Coding Challenge","Quiz","Project / Milestone","Theory (h)","Lab (h)","Total 45h","Total 60h"],
                   [6,26,40,34,40,34,16,34,9,8,9,9])
    for i,w in enumerate(weeks):
        wk,uref,topics,ex,coding,quiz,proj,th,lab=w
        uname=f"Unit {uref}: {units[uref-1][0]}" if isinstance(uref,int) else uref
        reading=reading_for(units[uref-1][0]) if isinstance(uref,int) else "Consolidation / project guidance + review checklist."
        vals=[wk,uname,topics,reading,ex,coding,quiz,proj,th,lab,round(th+lab,1),round(th+lab+1,1)]
        zb=ZEBRA if i%2==0 else WHITE
        for c,v in enumerate(vals,1):
            f=GROUP1 if c==2 else (HILITE if c in (5,6) else zb)
            cell(ws,r+i,c,v,fill=f,bold=(c==1),halign=('center' if c in (1,9,10,11,12) else 'left'))
        ws.row_dimensions[r+i].height=max(46,14*(1+max(len(topics),len(ex),len(reading))//42))
    tot=r+len(weeks)
    cell(ws,tot,2,"TOTAL",fill=HEADER,font=WHITEF,bold=True,halign='right')
    for c in [1]+list(range(3,9)): cell(ws,tot,c,"",fill=HEADER)
    cell(ws,tot,9,round(sum(w[7] for w in weeks),1),fill=HEADER,font=WHITEF,bold=True,halign='center')
    cell(ws,tot,10,round(sum(w[8] for w in weeks),1),fill=HEADER,font=WHITEF,bold=True,halign='center')
    cell(ws,tot,11,round(sum(w[7]+w[8] for w in weeks),1),fill=HEADER,font=WHITEF,bold=True,halign='center')
    cell(ws,tot,12,round(sum(w[7]+w[8]+1 for w in weeks),1),fill=HEADER,font=WHITEF,bold=True,halign='center')
    # Assessment & QB Plan
    ws=wb.create_sheet('Assessment & QB Plan')
    r=sheet_header(ws,f"Semester {sem_no} - Assessment & Question Bank Plan","Per-unit targets. Coding ratio 1:2 = 10 course + 20 question-bank questions. MCQ template provided; coding template pending.",
                   ["Unit","MCQ (total)","MCQ Beg","MCQ Int","MCQ Chal","Coding (course)","Coding (QB)","Total coding","Bloom focus"],[40,11,9,9,9,13,11,12,24])
    for i,(utitle,_,_) in enumerate(units):
        vals=[f"Unit {i+1}: {utitle}",10,4,3,3,10,20,30,"Remember/Understand" if i==0 else "Apply/Analyze/Create"]
        zb=ZEBRA if i%2==0 else WHITE
        for c,v in enumerate(vals,1):
            cell(ws,r+i,c,v,fill=(GROUP1 if c==1 else zb),bold=(c==1),halign=('left' if c in(1,9) else 'center'))
        ws.row_dimensions[r+i].height=24
    n=len(units); tt=r+n
    cell(ws,tt,1,"TOTAL",fill=HEADER,font=WHITEF,bold=True,halign='right')
    for c,v in enumerate([n*10,n*4,n*3,n*3,n*10,n*20,n*30],2): cell(ws,tt,c,v,fill=HEADER,font=WHITEF,bold=True,halign='center')
    cell(ws,tt,9,"All levels",fill=HEADER,font=WHITEF,bold=True,halign='left')
    # Question Taxonomy
    ws=wb.create_sheet('Question Taxonomy')
    r=sheet_header(ws,f"Semester {sem_no} - Question Taxonomy","Slugs for tagging every question (mirrors the Java template). subjects/topics/subTopics map to the MCQ template columns.",
                   ["subject","topic (unit)","subTopic","Unit"],[14,30,46,42])
    cur=r
    for ui,(utitle,_,topics) in enumerate(units,1):
        start=cur; tslug=slug(utitle)
        for ti,t in enumerate(topics):
            zb=ZEBRA if ti%2==0 else WHITE
            cell(ws,cur,1,'python',fill=zb); cell(ws,cur,3,slug(t),fill=zb); cell(ws,cur,4,f"Unit {ui}: {utitle}",fill=zb)
            ws.row_dimensions[cur].height=18; cur+=1
        ws.merge_cells(start_row=start,start_column=2,end_row=cur-1,end_column=2)
        cell(ws,start,2,tslug,fill=GROUP2,bold=True,valign='top')
    # Projects & Capstone
    ws=wb.create_sheet('Projects & Capstone')
    r=sheet_header(ws,f"Semester {sem_no} - Project Structure (4 levels)","Structural overview only; full briefs, expected output, and rubrics to be detailed in a later round.",
                   ["Level","Where it sits","Title / Theme","Concepts Covered"],[20,24,40,60])
    pr=[("Topic-level practice","Every session","Single-concept drills and code labs","One concept at a time; immediate application of the topic just taught."),
        ("Unit mini-project","End of major units","See per-unit themes in the Week-wise Plan","Combines the topics of one unit into a small, demoable artifact.")]
    pr+=sem_project
    pr+=[("Capstone (cross-semester)","Semester 2, built across Weeks 12-15",capstone_text[0],capstone_text[1])]
    for i,(a,b,c2,d) in enumerate(pr):
        zb=ZEBRA if i%2==0 else WHITE
        cell(ws,r+i,1,a,fill=GROUP1,bold=True); cell(ws,r+i,2,b,fill=zb); cell(ws,r+i,3,c2,fill=HILITE); cell(ws,r+i,4,d,fill=zb)
        ws.row_dimensions[r+i].height=max(34,15*(1+len(d)//58))
    # Templates
    ws=wb.create_sheet('Templates')
    r=sheet_header(ws,"Question Bank Templates","MCQ template matches Question Bank/Template/questions-mcq-template.xlsx. Coding template is pending your format (placeholder below).",
                   ["title","description","explanation","score","status","difficulty","bloomTaxonomy","tags","subjects","topics","subTopics","companies","option1","option2","option3","option4","answer"],
                   [16,22,22,7,10,10,14,12,12,12,12,12,12,12,12,12,8])
    samples=[["What does print() do?","Pick the best description.","print() writes its arguments to standard output.",5,"published","easy","remember","io,basics","python","introduction-to-programming","first-program","-","Reads input","Writes to output","Deletes a file","Defines a class",2]]
    for i,row in enumerate(samples):
        for c,v in enumerate(row,1): cell(ws,r+i,c,v,fill=ZEBRA,halign=('center' if c in(4,17) else 'left'))
        ws.row_dimensions[r+i].height=30
    pr2=r+len(samples)+1
    ws.merge_cells(start_row=pr2,start_column=1,end_row=pr2,end_column=17)
    cell(ws,pr2,1,"CODING-QUESTION TEMPLATE: PENDING. You will provide the format; this curriculum reserves 10 course + 20 question-bank coding questions per unit (1:2). Suggested columns to align once shared: title, problemStatement, constraints, sampleInput, sampleOutput, starterCode, referenceSolution, hiddenTests, difficulty, bloomTaxonomy, subjects, topics, subTopics, companies, score.",fill=HILITE,bold=True)
    ws.row_dimensions[pr2].height=64
    wb.save(path); return path

# ---------- narrative content ----------
OUT_S1=("By the end, students can: think computationally and design logic with pseudocode/flowcharts; use Python data types, operators, control flow, and loops; "
 "work fluently with strings, lists, tuples, sets, and dictionaries; write clean functions (including lambda, map/filter/reduce, nested functions, scope, and recursion); "
 "model simple objects with basic OOP; organize code into modules and packages and manage environments with venv/uv; read and write files using pathlib, glob, CSV, and JSON; "
 "handle errors with exceptions; and debug systematically.")
OUT_S2=("By the end, students can: explain how Python runs code (bytecode, __pycache__, imports); design systems with full OOP (encapsulation, abstraction, inheritance, polymorphism); "
 "write iterators, generators, decorators, and context managers; use the standard library confidently (random, hashlib/secrets, datetime, collections, itertools); "
 "test with pytest and enforce quality with type hints, linting, and pre-commit hooks; write concurrent code with asyncio, threads, and processes; build CLI apps with argparse/typer; "
 "interact with relational databases via a cursor; and structure, build, and ship an installable, documented package.")

MODEL_S1=("Semester 1 = 13 units. Data Structures is taught as three focused units (Strings; Lists and Tuples; Sets and Dictionaries), which keeps each unit weighty enough for a 30-question bank "
 "and maps cleanly to a 15-week term: 13 teaching units, then a semester-project week and a review/finals week. This stays within the 15-unit ceiling.")
MODEL_S2=("Semester 2 = 14 units. It opens with a Recap and Python Internals unit, splits advanced OOP into Encapsulation and Abstraction and a separate Inheritance and Polymorphism unit (Java-style), "
 "adds a Standard Library unit before Testing and a Database Interaction unit, and keeps Asynchronous Programming separate from a combined Multithreading and Multiprocessing unit. "
 "14 units fit a 15-week term with the capstone built across the final weeks and demoed in week 15, staying within the 15-unit ceiling.")

ADD_S1=("None required. The 13-unit Semester 1 list is complete. venv and uv are taught inside Modules and Packages, requirements/reproducible setups are marked optional there, "
 "logging is introduced inside Debugging, and File Handling now includes pathlib and glob.")
ADD_S2=("Now incorporated from earlier suggestions: a dedicated Standard Library unit and a Database Interaction unit. "
 "Still parked for a future revision if you want them: REST APIs and HTTP (requests), Regular Expressions (re), Design Patterns (SOLID, Factory, Strategy), and Performance profiling. "
 "Type hints and clean code remain folded into Code Quality and Pre-commit Hooks.")

ASSUME_S1=("1) No chapters; each item is a UNIT with topics. 2) Data Structures split into three units (Strings; Lists and Tuples; Sets and Dictionaries), giving 13 units total. "
 "3) Functions unit ordered with map/filter/reduce right after lambda, then built-ins (all, any, len, ...), nested functions, variable scope (taught here, after the Unit 2 variables material), and recursion. "
 "4) Modules and Packages: requirements/reproducible setups marked optional; mini-project is to build an installable package (Gold Price Index calculator) - exact theme to confirm. "
 "5) File Handling adds pathlib and glob. 6) Scope = structure + plan; actual questions and full project rubrics deferred. 7) Coding-question template pending from you.")
ASSUME_S2=("1) 14 units, opening with Recap and Python Internals (bytecode, __pycache__/.pyc, imports, framework intermediate files). "
 "2) Advanced OOP split into Encapsulation and Abstraction, and Inheritance and Polymorphism. "
 "3) Standard Library is its own unit before Testing; 'crypto' is read as the stdlib modules hashlib, secrets, and hmac (true crypto libraries are third-party). "
 "4) Database unit starts on SQLite, then PostgreSQL, with pgvector shown as a modern example. "
 "5) Multithreading and Multiprocessing kept as ONE unit so the total stays at 14, within the 15 cap. "
 "6) Capstone built across Weeks 12-15 and demoed in Week 15. 7) Coding-question template pending from you.")

CROSS_S1="Feeds directly into Semester 2: basic OOP becomes full OOP (encapsulation, abstraction, inheritance, polymorphism); files/with become context managers; modules/venv/uv become packaging and distribution; debugging/logging become testing and quality gates."
CROSS_S2="Builds directly on Semester 1: assumes comfort with functions, the data-structure units, files, basic OOP, modules, exceptions, and debugging. Every unit pushes toward production-grade, shippable Python and the cross-semester capstone."

SEMPROJ_S1=[("Unit mini-project","Unit 10 (Modules & Packages)","Build an installable package: Gold Price Index calculator","A reusable package other code can import; covers modules, packages, __init__.py, pip, venv/uv, and clean structure. Exact theme to confirm."),
 ("Semester project","Week 14","Command-Line Expense / Student Manager","Integrates Units 1-13: control flow, loops, the three data-structure units, functions, basic OOP, modules and packages, file persistence (CSV/JSON), exception handling, and debugging.")]
SEMPROJ_S2=[("Semester project","Rolls into the capstone","Professional, Tested, Packaged CLI Tool","Integrates Semester 2: full OOP, generators, decorators, context managers, the standard library, a pytest suite, type hints + pre-commit, async or threaded I/O, a typer CLI, a database backend, and packaging.")]

CAPSTONE=("Cross-Semester Capstone: Build, Test, and Ship a Real CLI Application (for example a Personal Finance Manager, Task and Habit Tracker, or Library Manager).",
 "Must combine BOTH semesters: fundamentals, data structures, functions, file handling, exceptions, modules and packages, basic and advanced OOP, the standard library, a database backend (SQLite or PostgreSQL), automated tests (pytest), a CLI (argparse/typer), debugging, type hints and pre-commit quality gates, and clean project organization packaged with pyproject/uv. Deliverable: a documented repository, a working installable command, and a demo.")

OUT='/Users/suman/Desktop/ByteXL/content/Curriculum'
p1=build(f"{OUT}/Python Curriculum - Semester 1 (Fundamentals).xlsx",1,"Python Fundamentals",
  "Think like a programmer, then build real Python: foundations to data structures, files, errors, and debugging.",
  SEM1_UNITS,SEM1_WEEKS,
  "Absolute beginners with no prior programming experience: first-year students, career-switchers, and self-learners.",
  "Basic computer literacy (files, browser, typing). No prior coding, math beyond school arithmetic, or terminal experience required.",
  OUT_S1,MODEL_S1,ADD_S1,ASSUME_S1,CROSS_S1,SEMPROJ_S1,CAPSTONE)
p2=build(f"{OUT}/Python Curriculum - Semester 2 (Advanced).xlsx",2,"Advanced Python Concepts",
  "From confident beginner to job-ready: internals, full OOP, the standard library, async, databases, testing, and shipping real tools.",
  SEM2_UNITS,SEM2_WEEKS,
  "Students who completed Python Fundamentals (Semester 1) or are comfortable with functions, the core data structures, files, basic OOP, modules, and errors.",
  "Completion of Semester 1 (Python Fundamentals) or equivalent: functions, data structures, files, basic OOP, modules/venv, exceptions, and debugging.",
  OUT_S2,MODEL_S2,ADD_S2,ASSUME_S2,CROSS_S2,SEMPROJ_S2,CAPSTONE)
print("Built:\n ",p1,"\n ",p2)
