import os, re
import openpyxl

BASE='/Users/suman/Desktop/ByteXL/content'
SEMS=[(1,'Python Fundamentals','Python Curriculum - Semester 1 (Fundamentals).xlsx'),
      (2,'Advanced Python Concepts','Python Curriculum - Semester 2 (Advanced).xlsx')]

def slug(s):
    s=s.split('(')[0]
    s=re.sub(r'[^a-zA-Z0-9 ]','',s).strip().lower()
    return re.sub(r'\s+','_',s)[:50]

def read_units(xlsx):
    wb=openpyxl.load_workbook(os.path.join(BASE,'Curriculum',xlsx))
    ws=wb['Curriculum (Unit-Topic)']
    units=[]; cur=None
    for r in range(5, ws.max_row+1):
        a=ws.cell(r,1).value; b=ws.cell(r,2).value
        if a:
            head=str(a).split('\n')
            title=head[0].split(':',1)[1].strip()
            summary=head[-1].strip() if len(head)>1 else ''
            cur={'title':title,'summary':summary,'topics':[]}; units.append(cur)
        if b and cur is not None:
            cur['topics'].append(str(b).strip())
    return units

STYLE=("Each lesson follows the house style: a standardized **Introduction** heading (no page-title H1), a story-led flow "
 "with real-world examples and situation-based questions under natural headings, a labelled diagram, and a **Final Takeaway**. "
 "Plain-English explanations are preferred; Python code is kept simple and interactive. No emojis, no em dashes.")

def unit_readme(sem_no, sem_name, n, unit):
    lines=[f"# Unit {n}: {unit['title']}", "", f"**Semester {sem_no}: {sem_name}**", ""]
    if unit['summary']: lines+=[unit['summary'], ""]
    lines+=["## Topics (teach in order)",""]
    for i,t in enumerate(unit['topics'],1):
        lines.append(f"{i}. {t}  — suggested file: `{i:02d}_{slug(t)}.md`".replace('—','->'))
    lines+=["", "## How each lesson is written", "", STYLE, "",
            "_Status: lesson content to be authored on demand. Semester 1 / Unit 1 is the worked reference._", ""]
    return "\n".join(lines)

def sem_index(sem_no, sem_name, units):
    lines=[f"# Semester {sem_no}: {sem_name}", "",
           f"{len(units)} units, taught in order over a 15-week term. See the curriculum workbook for the full week-wise plan and assessment design.", "",
           "## Units", "", "| # | Unit | Topics | Folder |", "|---|------|--------|--------|"]
    for i,u in enumerate(units,1):
        folder=f"Unit {i} - {u['title']}"
        lines.append(f"| {i} | {u['title']} | {len(u['topics'])} | [{folder}]({folder.replace(' ','%20')}/) |")
    lines+=["", "**Style:** professional, beginner-friendly, no emojis, no em dashes; standardized Introduction heading, narrative flow.", ""]
    return "\n".join(lines)

created=[]; skipped=[]
for sem_no, sem_name, xlsx in SEMS:
    units=read_units(xlsx)
    sem_dir=os.path.join(BASE, f"Semester {sem_no}")
    os.makedirs(sem_dir, exist_ok=True)
    with open(os.path.join(sem_dir,'README.md'),'w',encoding='utf-8') as f:
        f.write(sem_index(sem_no, sem_name, units))
    for i,u in enumerate(units,1):
        folder=os.path.join(sem_dir, f"Unit {i} - {u['title']}")
        # do not touch the already-authored Semester 1 / Unit 1
        if sem_no==1 and i==1:
            skipped.append(folder); continue
        os.makedirs(folder, exist_ok=True)
        rp=os.path.join(folder,'README.md')
        with open(rp,'w',encoding='utf-8') as f:
            f.write(unit_readme(sem_no, sem_name, i, u))
        created.append(folder)

print('created/updated unit folders:', len(created))
print('skipped (already authored):', skipped)
print('total Sem1 units:', len(read_units(SEMS[0][2])), '| total Sem2 units:', len(read_units(SEMS[1][2])))
