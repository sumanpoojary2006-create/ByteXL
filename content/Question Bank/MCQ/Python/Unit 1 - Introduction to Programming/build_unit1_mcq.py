"""Build the Unit 1 - Introduction to Programming MCQ workbook from the template.

Grounded strictly in content/Semester 1/Unit 1 - Introduction to Programming:
what programming is (clear/ordered/complete instructions, the literal machine) /
IPO model (input, processing, output; garbage in, garbage out; CPU/RAM/storage) /
computational thinking (decomposition, pattern recognition, abstraction, and the
fourth pillar algorithmic thinking) / problem-solving method (understand, plan,
carry out, look back; awkward inputs) / algorithms (5 properties; algorithm vs
program; tracing; efficiency) / pseudocode / flowcharts / why Python /
setting up Python and the first program (REPL vs script; print; input returns
text; storing a value is not output).

Writing rules (same bar as Unit 3):
* Simple English, short sentences.
* The idea is shown through a situation the student must reason about, not a
  definition to recall blindly.
* Each distractor maps to one specific, nameable misconception.
* Every question demands its own distinct reasoning step.

Set 1 (Q1-10)  -> tag "python - Set 1", titles 1.1.1 .. 1.1.10
Set 2 (Q11-40) -> tag "python - Set 2", titles 1.2.1 .. 1.2.30
Answer field is 1..4 (A=1, B=2, C=3, D=4).
"""

import os
import shutil
from collections import Counter

import openpyxl

TEMPLATE = "content/Question Bank/Template/questions-mcq-template.xlsx"
OUT_DIR = "content/Question Bank/MCQ/Python/Unit 1 - Introduction to Programming"
OUT_FILE = os.path.join(OUT_DIR, "Unit 1 - Introduction to Programming - MCQ.xlsx")
SHEET_NAME = "Python - MCQ - 1"

L = {"A": 1, "B": 2, "C": 3, "D": 4}
SCORE = {"easy": 5, "medium": 8, "hard": 10}

Q = []


def add(setn, diff, bloom, sub, desc, expl, opts, ans):
    Q.append((setn, diff, bloom, sub, desc, expl, opts, ans))


# =====================  SET 1  (Q1-Q10)  =====================

# Q1 - complete instructions (a missing step)
add(1, "medium", "analyze", "what-is-programming",
    "A student writes steps for a robot to make toast: (1) put the bread in the toaster, (2) press the lever, (3) spread butter on the toast. The robot burns itself trying to butter bread that is still inside the running toaster.\n\n"
    "Which quality of good instructions is missing?",
    "Good instructions must be unambiguous, ordered, and complete. Here a required step — taking the toast out before buttering it — was never written, so the steps are not complete. A literal machine will not add a step you left out.",
    ("It is not unambiguous — step 2 could mean two different things",
     "It is not complete — a needed step (take the toast out) is missing",
     "It has too much detail the robot does not need",
     "It is written in English instead of Python"),
    "B")

# Q2 - IPO: classify the calculation stage
add(1, "easy", "understand", "inputs-processing-outputs",
    "A fitness band reads your steps from a sensor, works out how many calories you burned, and shows that number on its screen.\n\n"
    "In the IPO model, what is \"working out the calories\"?",
    "IPO stands for Input, Processing, Output. Reading the sensor is input and showing the number is output. Working out the calories is a calculation done on the input, which is the processing stage.",
    ("Input, because the steps came from a sensor",
     "Output, because a number appears on the screen",
     "Storage, because the band keeps your history",
     "Processing, because it is a calculation done on the input"),
    "D")

# Q3 - computational thinking: pattern recognition
add(1, "medium", "understand", "computational-thinking",
    "While planning a college fest, an organiser notices that printing certificates, printing fee receipts, and printing ID cards are all the same \"one record per person\" job — so one solution can be reused for all three.\n\n"
    "Which pillar of computational thinking is this?",
    "Noticing that different tasks share the same underlying shape, so one solution can be reused, is pattern recognition. Decomposition is splitting a big job into smaller ones, and abstraction is dropping details that do not matter.",
    ("Pattern recognition — spotting the same shape of task repeating",
     "Decomposition — splitting a big task into smaller ones",
     "Abstraction — ignoring the details that do not matter",
     "Debugging — fixing an error after it appears"),
    "A")

# Q4 - algorithm property: definiteness
add(1, "hard", "analyze", "algorithms",
    "An algorithm for cooking includes the step \"season the dish somehow\". A classmate says this cannot be a proper algorithm step.\n\n"
    "Which property does that step break?",
    "Definiteness means every step has exactly one meaning. \"Somehow\" can be read in many ways, so two people could do it differently — that breaks definiteness. It is not about the step running forever (finiteness) or lacking data (input).",
    ("Finiteness — the step might run forever",
     "Output — the step produces no result",
     "Definiteness — \"somehow\" can be read in more than one way",
     "Input — the step needs data it does not have"),
    "C")

# Q5 - first program: storing a value prints nothing
add(1, "medium", "analyze", "setting-up-python-first-program",
    "A beginner runs this program and says \"only two lines appear, so something must be broken\":\n\n"
    "```python\nprint(\"Start\")\ntotal = 40 + 2\nprint(\"Total is\", total)\n```\n\n"
    "How many lines appear, and is anything broken?",
    "`print(\"Start\")` shows one line, and `print(\"Total is\", total)` shows `Total is 42`. The middle line only stores a value, which is processing, not output, so it shows nothing. Two lines is exactly right — nothing is broken.",
    ("Three lines — every line prints something",
     "Two lines — storing a value shows nothing, so nothing is broken",
     "One line — only the last print runs",
     "Nothing — the program has an error"),
    "B")

# Q6 - flowchart shape for a decision
add(1, "easy", "remember", "flowcharts",
    "In a flowchart, one shape marks a yes/no question where the path splits into two branches.\n\n"
    "Which shape is it?",
    "A diamond marks a decision — a yes/no question that splits the path. An oval starts or ends the chart, a rectangle is a plain process step, and a parallelogram is for input or output.",
    ("Oval",
     "Rectangle",
     "Parallelogram",
     "Diamond"),
    "D")

# Q7 - reading/tracing pseudocode
add(1, "medium", "apply", "pseudocode",
    "Read this pseudocode:\n\n"
    "```\nREAD mark1, mark2, mark3\nSET average = (mark1 + mark2 + mark3) / 3\nIF average >= 40 THEN\n    PRINT \"Pass\"\nELSE\n    PRINT \"Fail\"\nEND IF\n```\n\n"
    "For marks 30, 35, and 30, what does it print?",
    "Add the marks: 30 + 35 + 30 = 95. Divide by 3 to get about 31.7. Since 31.7 is below 40, the `IF` is false and the `ELSE` runs, printing `Fail`. Pseudocode can be traced by hand exactly like this.",
    ("Fail — the average is about 31.7, which is below 40",
     "Pass — the average is above 40",
     "Nothing — pseudocode can never be traced",
     "Both Pass and Fail"),
    "A")

# Q8 - why Python reads so cleanly
add(1, "medium", "understand", "why-python",
    "A beginner is surprised that Python code to greet three friends reads almost like an English sentence, while the same task in many older languages first needs type declarations, semicolons, and curly braces.\n\n"
    "What makes Python read so cleanly?",
    "Python's designers chose English-like keywords, very little boilerplate, and indentation that is part of the language itself. This keeps the code and the idea behind it looking almost the same, instead of burying the idea under punctuation.",
    ("It has no keywords at all, so there is nothing extra to read",
     "It must be compiled first, and that step tidies the code",
     "It uses English-like keywords, little boilerplate, and indentation as part of the language",
     "It only works for very short programs, so the code stays small"),
    "C")

# Q9 - input() always returns text
add(1, "hard", "analyze", "setting-up-python-first-program",
    "A program asks for a number using `input()`, and the user types 25. A student expects Python to store the number 25 so it can do maths with it right away.\n\n"
    "What does Python actually store, and why does it matter?",
    "`input()` always hands back plain text, so Python stores the text `\"25\"`, not the number 25. Before doing any maths you must convert it to a number first. This is a detail the next unit builds on directly.",
    ("The text \"25\" — input() always returns text, so you must convert it before doing maths",
     "The number 25 — input() detects numbers on its own",
     "Nothing — input() only reads, it does not store",
     "The number 25.0 — input() turns everything into a decimal"),
    "A")

# Q10 - problem-solving: pick the boundary test case
add(1, "hard", "analyze", "problem-solving-approach",
    "A pass/fail rule says \"an average of 40 or more passes\". A careful student wants the one test case most likely to catch a mistake where someone wrote `>` instead of `>=`.\n\n"
    "Which set of marks should they test?",
    "The `>` vs `>=` mistake only shows up right at the pass mark. Marks of 40, 40, 40 give an average of exactly 40, which must pass under \"40 or more\" but would wrongly fail if `>` were used. Clearly-passing or clearly-failing marks cannot reveal that mistake.",
    ("90, 85, 95 — all clearly passing",
     "10, 5, 0 — all clearly failing",
     "40, 40, 40 — an average sitting exactly on the pass mark",
     "1000, 100, 100 — an unusually large value"),
    "C")


# =====================  SET 2  (Q11-Q40)  =====================

# Q11 - using vs writing software
add(2, "easy", "understand", "what-is-programming",
    "Consider two actions: tapping \"Add to Cart\" on a shopping app, and writing the logic that runs the moment that button is tapped.\n\n"
    "Which one is writing software?",
    "Using software means following choices someone already built, like tapping a button. Writing software means deciding what those steps should be — writing the logic behind the button. So writing the logic is programming.",
    ("Tapping the button, because you are using the feature",
     "Both are the same activity",
     "Neither one involves programming",
     "Writing the logic behind the button, because you decide the steps"),
    "D")

# Q12 - the literal machine
add(2, "medium", "understand", "what-is-programming",
    "A helper follows instructions exactly and never fills in a step you leave out. You tell it: spread the butter, put the second slice on top, serve — but you forget the step \"take out two slices of bread first\".\n\n"
    "What does this literal helper do?",
    "A computer, like the literal helper, follows instructions exactly as written and never adds a missing step. With no bread taken out, it tries to spread butter on nothing. It does not guess your intention or fix the gap for you.",
    ("It guesses you meant to get bread and carries on",
     "It tries to spread butter with no bread, because that step was never given",
     "It stops and asks you for the missing step",
     "It fixes the order for you automatically"),
    "B")

# Q13 - garbage in, garbage out
add(2, "hard", "analyze", "inputs-processing-outputs",
    "An exam portal is given a student's mark as the word \"ninety\" instead of the number 90. The processing logic is flawless, yet the final result is wrong.\n\n"
    "Which idea does this show?",
    "\"Garbage in, garbage out\" means that even perfect processing cannot rescue bad input — a wrong or badly-shaped input leads to a wrong output. This is exactly why careful programs check their input before trusting it.",
    ("Garbage in, garbage out — bad input gives a wrong result even with perfect processing",
     "Finiteness — the program runs forever",
     "Abstraction — ignoring details that do not matter",
     "Output first — the result should be designed before the input"),
    "A")

# Q14 - RAM vs storage vs CPU
add(2, "medium", "understand", "inputs-processing-outputs",
    "While a program runs, it holds the PIN you just typed, but loses it the moment the program closes.\n\n"
    "Which part of the computer holds data that disappears like this?",
    "RAM is short-term memory: it holds the data a program is actively using but is cleared when the program closes. Storage keeps files even with the power off, and the CPU does the calculations rather than holding working data.",
    ("The CPU, the part that does the calculations",
     "Storage, the long-term memory",
     "RAM, the short-term memory that is cleared when the program closes",
     "The input device, such as the keypad"),
    "C")

# Q15 - decomposition
add(2, "easy", "understand", "computational-thinking",
    "Nobody writes a whole banking app in one go. It is split into login, balance, transfer, and statements, and each of those is split further until every piece is small enough to build.\n\n"
    "Which pillar of computational thinking is this?",
    "Breaking a big, overwhelming problem into smaller, manageable sub-problems is decomposition. Abstraction is dropping unimportant detail, and pattern recognition is spotting repeated shapes of task.",
    ("Abstraction",
     "Pattern recognition",
     "Efficiency",
     "Decomposition"),
    "D")

# Q16 - algorithm vs program
add(2, "medium", "understand", "algorithms",
    "A friend claims the plain-English steps for finding the largest of three numbers and the Python version of it are two completely unrelated things.\n\n"
    "Which statement is correct?",
    "An algorithm is the language-independent plan written in plain steps. A program is that same plan written in one specific language. So the English steps are the algorithm, and the Python version is one program that implements it.",
    ("They are unrelated — the English steps are not an algorithm",
     "The English steps are the algorithm; the Python version is one program of that same algorithm",
     "Only the Python version counts as an algorithm",
     "An algorithm must always be written in Python"),
    "B")

# Q17 - finiteness
add(2, "hard", "analyze", "algorithms",
    "A set of steps keeps repeating and has no way to ever stop.\n\n"
    "Why is this not a true algorithm?",
    "Every algorithm must be finite — guaranteed to stop after a limited number of steps. Steps that could run forever break finiteness, which is exactly why infinite loops are treated as bugs. The steps here may be perfectly clear, so definiteness is not the issue.",
    ("It breaks finiteness — an algorithm must be guaranteed to stop",
     "It breaks definiteness — the steps are unclear",
     "It breaks input — it needs no data to start",
     "It is fine — algorithms are allowed to run forever"),
    "A")

# Q18 - problem-solving: the look-back stage
add(2, "medium", "understand", "problem-solving-approach",
    "After building a program, a developer tests it with unusual inputs like all zeros and a value sitting exactly on the boundary, then improves it based on what they find.\n\n"
    "Which stage of the four-stage problem-solving method is this?",
    "The four stages are Understand, Plan, Carry Out, and Look Back. Testing with tricky inputs and improving the result is the Look Back stage — the step many beginners skip, and the one that catches subtle bugs.",
    ("Understand",
     "Plan",
     "Look Back — testing with tricky inputs and improving",
     "Carry Out"),
    "C")

# Q19 - flowchart backward arrow is a loop
add(2, "medium", "understand", "flowcharts",
    "In a login flowchart, an arrow points from \"wrong password\" back up to \"read password\" so the user can try again.\n\n"
    "What does a backward arrow represent?",
    "A backward arrow sends the flow to an earlier step so it runs again — that is repetition, which is a loop. It is the same idea you later write as a repeating loop in code.",
    ("The end of the program",
     "An input step",
     "A decision that always fails",
     "Repetition — the same steps run again, which is a loop"),
    "D")

# Q20 - why Python: indentation is part of the language
add(2, "medium", "analyze", "why-python",
    "A student removes all the indentation from working Python code to \"tidy it up\", and the code stops running.\n\n"
    "Why did that happen?",
    "In Python, indentation is part of the language, not just decoration — it shows which lines belong together. Remove it and Python can no longer tell the structure, so the code breaks. Python uses indentation where other languages use braces.",
    ("Python needs a semicolon at the end of each line, not indentation",
     "In Python, indentation is part of the language, not just decoration",
     "Indentation only matters in older languages, not in Python",
     "The code was already broken before the indentation was removed"),
    "B")

# Q21 - pseudocode is for humans
add(2, "medium", "understand", "pseudocode",
    "A teammate asks why you sketched pseudocode on the whiteboard instead of writing real code.\n\n"
    "What is the best reason?",
    "Pseudocode captures an algorithm's logic in plain words plus a few keywords, written for humans to reason about — not for a computer to run. It lets you think about what should happen without fighting a language's exact syntax.",
    ("Pseudocode captures the logic in plain words for humans to reason about, without fussy syntax",
     "Pseudocode runs faster than real code on a computer",
     "Pseudocode is the only form a computer can actually read",
     "Pseudocode removes the need to ever write real code"),
    "A")

# Q22 - abstraction: dropping an important detail breaks it
add(2, "hard", "analyze", "computational-thinking",
    "Abstraction means keeping the details that matter and dropping the rest. At a registration desk you keep the name and the ticket number, and ignore the colour of each student's shirt.\n\n"
    "What happens if you also abstract away the ticket number?",
    "Abstraction removes irrelevant detail, never important detail. The ticket number is essential to registration, so dropping it is dropping something that matters — and the solution quietly breaks. Shirt colour was safe to drop; the ticket number was not.",
    ("Nothing — abstraction lets you drop any detail safely",
     "The desk works faster because there is less data",
     "The solution breaks — the ticket number was an important detail, not noise",
     "It turns into a different pillar called decomposition"),
    "C")

# Q23 - which line shows nothing (storing is not output)
add(2, "hard", "analyze", "setting-up-python-first-program",
    "Look at this program:\n\n"
    "```python\nprint(\"Welcome!\")\nscore = 10 + 5\nprint(\"Your score is\", score)\n```\n\n"
    "Which line puts nothing on the screen, and why?",
    "The first line prints `Welcome!` and the third prints `Your score is 15`. The middle line works out 15 and stores it in `score`, but storing a value is processing, not output — only `print` puts something on screen. So the middle line shows nothing.",
    ("The first line — text in quotes is never printed",
     "The third line — score has no value yet",
     "None — all three lines print something",
     "The middle line — storing a value is processing, not output"),
    "D")

# Q24 - two correct algorithms, different speed (efficiency)
add(2, "hard", "analyze", "algorithms",
    "To find \"Ravi\" among 10,000 contacts, you could check every name from the top, or jump straight to the R names and scan only those. Both always give the right answer.\n\n"
    "What is the main difference between them?",
    "Both approaches are correct, definite, and finite — they always find Ravi. The difference is efficiency: how quickly they finish. Jumping to the R names checks far fewer names, so it is much faster on a big list.",
    ("Only the second one is a correct algorithm",
     "They are equally correct, but they differ in speed (efficiency)",
     "The first one may never finish",
     "The second one sometimes skips Ravi"),
    "B")

# Q25 - input can be a scheduled trigger
add(2, "medium", "understand", "inputs-processing-outputs",
    "A backup app wakes up on its own at 2 AM every night and runs, with no person typing anything at all.\n\n"
    "In the IPO model, what started it?",
    "Input is not only typing on a keyboard — it can be a tap, a file, a sensor reading, a network message, or a scheduled trigger. The 2 AM timer is a scheduled trigger, which is a form of input that set the program running.",
    ("Input — a scheduled trigger is a form of input",
     "Output — the backup is the result",
     "Processing — waking up is a calculation",
     "Nothing — a program cannot start without a person"),
    "A")

# Q26 - flowchart shape for input/output
add(2, "easy", "remember", "flowcharts",
    "On a flowchart, one shape is used for reading data in or showing a result out.\n\n"
    "Which shape is it?",
    "A parallelogram stands for input or output — reading data in or showing a result out. A diamond is a decision, an oval starts or ends the chart, and a rectangle is a plain process step.",
    ("Diamond",
     "Oval",
     "Parallelogram",
     "Rectangle"),
    "C")

# Q27 - understand before building
add(2, "medium", "understand", "problem-solving-approach",
    "A student cannot explain their plan for a task in plain language, yet keeps typing code and getting more confused.\n\n"
    "What does the problem-solving method say they should do?",
    "The rule is: understand first, plan second, build third. If you cannot explain your plan in plain language, the problem is not yet understood — the signal is to return to the Understand stage, not to type harder.",
    ("Keep typing faster until it works",
     "Skip straight to testing the program",
     "Add more comments to the broken code",
     "Go back and understand the problem before building"),
    "D")

# Q28 - why Python: batteries included / PyPI
add(2, "medium", "understand", "why-python",
    "A student needs to read data from the internet and does not want to build that ability from scratch. Python's design makes this easy.\n\n"
    "Which idea is this?",
    "Python ships with a large set of ready-made tools (\"batteries included\"), and PyPI offers hundreds of thousands of free packages. So someone has usually built the hard part already, and you assemble the pieces instead of starting from zero.",
    ("Indentation is part of the language",
     "\"Batteries included\" plus PyPI — ready-made tools and packages you can reuse",
     "Python must be compiled before it can run",
     "Python only suits small learning programs"),
    "B")

# Q29 - definiteness: which step is clear enough
add(2, "medium", "analyze", "algorithms",
    "Only one of the steps below is written clearly enough — with exactly one meaning — to belong in a real algorithm.\n\n"
    "Which step is it?",
    "A real algorithm step must be definite: precise, with exactly one meaning. \"Compare each pair of neighbours and swap them if they are out of order\" can be read only one way. \"Sort somehow\", \"make it nice\", and \"do the usual thing\" are all too vague to act on.",
    ("Compare each pair of neighbours and swap them if they are out of order",
     "Sort the list somehow",
     "Make the list nice",
     "Do the usual thing to the numbers"),
    "A")

# Q30 - pseudocode: BEGIN and END
add(2, "easy", "remember", "pseudocode",
    "In pseudocode, the keywords BEGIN and END are used at the top and bottom of the logic.\n\n"
    "What is their job?",
    "BEGIN and END are boundary keywords: they mark where the logic starts and stops. READ and PRINT handle input and output, FOR and WHILE handle repetition, and IF handles a decision.",
    ("They read input and show output",
     "They repeat a set of steps",
     "They mark where the logic starts and stops",
     "They make a yes/no decision"),
    "C")

# Q31 - order of instructions matters
add(2, "medium", "analyze", "what-is-programming",
    "A tea recipe lists the steps as: (1) pour into a cup, (2) boil the water, (3) add the tea powder. Followed exactly, it gives a poor result.\n\n"
    "What is wrong with these instructions?",
    "Good instructions must be ordered — the sequence matters. Here the steps are simply in the wrong order: you cannot pour the tea into a cup before the water is even boiled. No step is missing or ambiguous; they are just out of sequence.",
    ("A required step is missing",
     "A step is ambiguous",
     "There are too many steps",
     "The steps are in the wrong order — you cannot pour before boiling"),
    "D")

# Q32 - REPL for a quick throwaway answer
add(2, "easy", "understand", "setting-up-python-first-program",
    "You just want a quick, throwaway answer to \"what is 1234 times 5678?\" right now, with nothing to save.\n\n"
    "Which tool fits best?",
    "The REPL (Python's interactive shell) is made for quick, throwaway questions — you type the calculation and the answer appears at once, with nothing to save. A script is better when you want to keep and rerun something.",
    ("A saved script, because it keeps a record",
     "The REPL, which is made for quick, throwaway answers",
     "A flowchart, to draw the calculation",
     "Pseudocode, to plan the calculation"),
    "B")

# Q33 - the fourth pillar: algorithmic thinking
add(2, "medium", "understand", "computational-thinking",
    "After breaking a problem down, spotting its repeating pattern, and ignoring the noise, one last step turns it into a precise, ordered sequence of steps.\n\n"
    "What is that step called?",
    "Turning your decomposed, pattern-spotted, abstracted problem into a precise, ordered sequence of steps is algorithmic thinking — often named as the fourth pillar. It leads directly to writing an actual algorithm.",
    ("Algorithmic thinking — turning the plan into precise, ordered steps",
     "Abstraction",
     "Decomposition",
     "Pattern recognition"),
    "A")

# Q34 - why Python: interpreted, runs almost immediately
add(2, "medium", "understand", "why-python",
    "Compared with a typical compiled language such as C or Java, how does Python usually get from written code to running?\n\n"
    "Which statement is correct?",
    "Python runs your code almost immediately, without a separate compile step first. A typical compiled language is usually turned into machine code in a separate step before it can run.",
    ("It must always be compiled in a separate step first",
     "It cannot run until it is converted into another language",
     "It runs your code almost immediately, without a separate compile step",
     "It runs only inside a web browser"),
    "C")

# Q35 - first program: forgetting quotes
add(2, "medium", "analyze", "setting-up-python-first-program",
    "A beginner writes `print(Welcome)` without quotation marks around the word, and gets an error.\n\n"
    "Why?",
    "Text you want to show must be wrapped in quotes. Without them, Python treats `Welcome` as a name or command to look up rather than as a message to display, so it errors. Writing `print(\"Welcome\")` fixes it.",
    ("print always needs a number, not text",
     "Without quotes, Python treats Welcome as a name/command, not as text to show",
     "print cannot display a single word",
     "The line needs a semicolon at the end"),
    "B")

# Q36 - what is programming: describe once, run many times
add(2, "medium", "understand", "what-is-programming",
    "A program prints 5,000 fee receipts in under a second, every one correct, while a tired clerk would take days by hand and make mistakes.\n\n"
    "What did the programmer do that the clerk did not?",
    "The programmer stopped doing the work and started describing it: capturing the task as a clear, ordered set of instructions the computer can carry out again and again. The computer supplies the speed; the programmer supplies the thinking — no advanced maths required.",
    ("Worked faster by hand",
     "Removed the need for any thinking",
     "Used advanced mathematics",
     "Described the task once as clear steps for the computer to repeat"),
    "D")

# Q37 - IPO: identify the processing step
add(2, "easy", "understand", "inputs-processing-outputs",
    "You are designing a shopping-bill total. Your plan: ask for the price and quantity, multiply them together, then show the total.\n\n"
    "Which part is the processing step?",
    "Asking for the price and quantity is input, and showing the total is output. Multiplying price by quantity is the calculation in the middle — the processing step. The same processing works for any values the user enters.",
    ("Multiplying the price by the quantity",
     "Asking for the price and quantity",
     "Showing the total",
     "Saving the bill to a file"),
    "A")

# Q38 - trace an algorithm and know why it finishes
add(2, "hard", "apply", "algorithms",
    "Trace this algorithm for the numbers 12, 27, and 19: assume the first number is the largest; if the second is bigger, it becomes the largest; if the third is bigger, it becomes the largest.\n\n"
    "What is the final answer, and why is it guaranteed to finish?",
    "Start with 12. Is 27 bigger? Yes, so the largest becomes 27. Is 19 bigger than 27? No, so it stays 27. The answer is 27. It always finishes because it makes a fixed number of comparisons, no matter what the numbers are — that is its finiteness.",
    ("19 — the last number always wins",
     "27 — and it finishes because it makes a fixed number of comparisons",
     "12 — the first number is always kept",
     "It never finishes — the steps can loop forever"),
    "B")

# Q39 - flowchart: label every decision branch
add(2, "medium", "analyze", "flowcharts",
    "A flowchart has a decision diamond \"Average 40 or more?\" but only one arrow leaves it, and that arrow has no label.\n\n"
    "What is the problem?",
    "Every decision diamond must have a labelled exit for each outcome (such as yes and no). With one unlabelled arrow, the reader is left guessing what happens for each answer. The diamond itself is the right shape for a decision.",
    ("Diamonds should never be used for decisions",
     "A decision must have at least three exits",
     "Each branch must be labelled (such as yes and no), or the reader is left guessing",
     "Decisions cannot appear in a flowchart at all"),
    "C")

# Q40 - myths about programming
add(2, "medium", "understand", "what-is-programming",
    "A capable student avoids programming, believing \"you must be a maths genius, and you must memorise everything\".\n\n"
    "Based on this unit, what is the honest response?",
    "Neither belief is true. For most programming, school arithmetic is plenty — the real skill is logical, step-by-step thinking. And you do not memorise everything: even professionals look things up constantly. Programming is a skill built through practice.",
    ("True — only maths geniuses can program",
     "True — you must memorise all the syntax",
     "Half true — you must memorise everything but need no maths",
     "Neither is required — programming is a skill built by practice, and even professionals look things up"),
    "D")


# =====================  Build workbook  =====================

assert len(Q) == 40, f"expected 40 questions, got {len(Q)}"

os.makedirs(OUT_DIR, exist_ok=True)
shutil.copyfile(TEMPLATE, OUT_FILE)

wb = openpyxl.load_workbook(OUT_FILE)
ws = wb.active
ws.title = SHEET_NAME
ws.delete_rows(2, ws.max_row)  # keep header, drop sample rows

set_counter = {1: 0, 2: 0}
letters = []
answer_tally = Counter()

for setn, diff, bloom, sub, desc, expl, opts, ans in Q:
    set_counter[setn] += 1
    title = f"Python - MCQ - 1.{setn}.{set_counter[setn]}"
    tag = f"python - Set {setn}"
    answer_tally[ans] += 1
    letters.append(ans)
    ws.append([
        title, desc, expl, SCORE[diff], "published", diff, bloom,
        tag, "python", "introduction-to-programming", sub, None,
        opts[0], opts[1], opts[2], opts[3], L[ans],
    ])

wb.save(OUT_FILE)

# =====================  Validation  =====================
print("Saved:", OUT_FILE)
print("Total questions:", len(Q))
print("Answer distribution:", dict(sorted(answer_tally.items())))

run = max_run = 1
for a, b in zip(letters, letters[1:]):
    run = run + 1 if a == b else 1
    max_run = max(max_run, run)
print("Max consecutive same letter:", max_run)

assert all(answer_tally[x] == 10 for x in "ABCD"), answer_tally
assert max_run <= 2, "answer letter repeats more than twice consecutively"

chk = openpyxl.load_workbook(OUT_FILE)
cs = chk[SHEET_NAME]
rows = list(cs.iter_rows(values_only=True))
assert rows[0] == ('title', 'description', 'explanation', 'score', 'status',
                   'difficulty', 'bloomTaxonomy', 'tags', 'subjects', 'topics',
                   'subTopics', 'companies', 'option1', 'option2', 'option3',
                   'option4', 'answer'), rows[0]
assert len(rows) == 41, len(rows)
for r in rows[1:]:
    assert len(set(r[12:16])) == 4, f"duplicate options in {r[0]}"
    assert r[12:16][r[16] - 1], f"empty correct option in {r[0]}"
print("Difficulty mix:", dict(Counter(r[5] for r in rows[1:])))
print("Set 1:", sum(1 for r in rows[1:] if r[7] == 'python - Set 1'),
      "| Set 2:", sum(1 for r in rows[1:] if r[7] == 'python - Set 2'))
print("Validation passed.")
