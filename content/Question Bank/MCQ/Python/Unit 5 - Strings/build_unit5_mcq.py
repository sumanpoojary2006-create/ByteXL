"""Build the Unit 5 - Strings MCQ workbook from the shared template.

Grounded strictly in content/Semester 1/Unit 5 - Strings:
what a string is (quotes, ordered sequence, len counts spaces, empty string) /
indexing and slicing (index from 0 and -1, IndexError, start:stop with stop
excluded, [::-1], [-4:]) / immutability (item assignment TypeError; build a new
string; id changes) / common methods (upper/lower/title, strip, replace all,
isdigit/isalpha booleans; methods return a NEW string so you must reassign) /
split and join (split -> list, split() on spaces, sep.join(list)) / searching
(in, find -> index or -1, count, startswith/endswith, case-sensitive) /
f-string formatting (:.2f, :, :.1%, width and <>^ alignment) / escape sequences
(\\n, \\t, \\", \\\\, raw r"...", triple-quoted) / practical text processing
(chaining strip/lower/replace, split then unpack, word/char counts).

Writing rules (same bar as earlier units):
* Simple English, short sentences.
* Trace-values are chosen so a wrong reasoning path gives a different answer.
* Each distractor maps to one specific, nameable misconception.
* Every question demands its own distinct reasoning step.
* Set 1 (Q1-10) covers all nine lessons.

Set 1 (Q1-10)  -> tag "python - Set 1", titles 5.1.1 .. 5.1.10
Set 2 (Q11-40) -> tag "python - Set 2", titles 5.2.1 .. 5.2.30
Answer field is 1..4 (A=1, B=2, C=3, D=4).
"""

import os
import shutil
from collections import Counter

import openpyxl

TEMPLATE = "content/Question Bank/Template/questions-mcq-template.xlsx"
OUT_DIR = "content/Question Bank/MCQ/Python/Unit 5 - Strings"
OUT_FILE = os.path.join(OUT_DIR, "Unit 5 - Strings - MCQ.xlsx")
SHEET_NAME = "Python - MCQ - 5"

L = {"A": 1, "B": 2, "C": 3, "D": 4}
SCORE = {"easy": 5, "medium": 8, "hard": 10}

Q = []


def add(setn, diff, bloom, sub, desc, expl, opts, ans):
    Q.append((setn, diff, bloom, sub, desc, expl, opts, ans))


# =====================  SET 1  (Q1-Q10)  =====================

# Q1 - len counts spaces
add(1, "easy", "understand", "what-is-a-string",
    "Meera's app shows a live character counter as she types. She enters the text \"Sea Shells\", which has one space in the middle:\n\n"
    "```python\ntext = \"Sea Shells\"\nprint(len(text))\n```\n\n"
    "What number does the counter show?",
    "`len()` counts every character in the string, and a space is a character too. \"Sea Shells\" is S, e, a, space, S, h, e, l, l, s — 10 characters in all.",
    ("10 — len counts every character, and the space counts too",
     "9 — spaces are not counted",
     "2 — it counts the words",
     "8 — it counts only the letters"),
    "A")

# Q2 - indexing: index 0 and -1
add(1, "medium", "analyze", "indexing-and-slicing",
    "What does this print?\n\n"
    "```python\nword = \"Python\"\nprint(word[0], word[-1])\n```",
    "Indexing counts from 0 at the front, so `word[0]` is the first letter, `P`. Negative indexing counts from the back, so `word[-1]` is the last letter, `n`. The output is `P n`.",
    ("y n — index 0 is the second letter",
     "P o — -1 is the second-to-last letter",
     "P n — index 0 is the first letter, -1 is the last",
     "An error — you cannot use -1"),
    "C")

# Q3 - immutability: item assignment raises TypeError
add(1, "hard", "analyze", "string-immutability",
    "What happens when this runs?\n\n"
    "```python\nword = \"Python\"\nword[0] = \"J\"\nprint(word)\n```",
    "Strings are immutable — you cannot change a character in place. `word[0] = \"J\"` raises a `TypeError` saying the string does not support item assignment. To change text you build a new string instead.",
    ("It prints Jython — the first letter is replaced",
     "It raises a TypeError — strings cannot be changed in place",
     "It prints Python — the change is ignored silently",
     "It prints J — only the changed letter shows"),
    "B")

# Q4 - method result not stored (classic bug)
add(1, "medium", "analyze", "common-string-methods",
    "What is printed, and why?\n\n"
    "```python\nname = \"asha\"\nname.upper()\nprint(name)\n```",
    "`upper()` does not change the original string — it returns a new one. Here that new string `\"ASHA\"` is never stored, so it is thrown away and `name` is still `\"asha\"`. The fix is `name = name.upper()`.",
    ("ASHA — upper() changes name in place",
     "An error — upper() needs an argument",
     "A — only the first letter changes",
     "asha — upper() returns a new string that was never stored"),
    "D")

# Q5 - split produces a list
add(1, "medium", "analyze", "splitting-and-joining",
    "What does this print?\n\n"
    "```python\ndata = \"Asha,20,Pune\"\nparts = data.split(\",\")\nprint(parts)\n```",
    "`split(\",\")` cuts the string at each comma and hands back a list of the pieces. So `parts` becomes the list `['Asha', '20', 'Pune']`.",
    ("['Asha', '20', 'Pune'] — split cuts at each comma into a list",
     "Asha 20 Pune — one string with spaces",
     "'Asha,20,Pune' — split does nothing without help",
     "An error — you cannot split a string"),
    "A")

# Q6 - find returns -1 when absent
add(1, "medium", "analyze", "searching-within-strings",
    "What does `find` return here?\n\n"
    "```python\nemail = \"asha@email.com\"\nprint(email.find(\"z\"))\n```",
    "`find` returns the index where the text first appears, or `-1` if it is not there at all. There is no `z` in the email, so it returns `-1`.",
    ("0 — z is treated as the start",
     "An error — the letter is missing",
     "-1 — find returns -1 when the text is not there",
     "None — find has no result to give"),
    "C")

# Q7 - f-string .2f
add(1, "medium", "apply", "string-formatting-and-fstrings",
    "What does this print?\n\n"
    "```python\nprice = 49.5\nprint(f\"Total: {price:.2f}\")\n```",
    "The `:.2f` after the colon means \"show this number with exactly two decimal places\". 49.5 becomes `49.50`, so it prints `Total: 49.50`.",
    ("Total: 49.5",
     "Total: 49",
     "Total: 50.00 — it rounds up",
     "Total: 49.50 — .2f shows two decimal places"),
    "D")

# Q8 - escape sequence \n
add(1, "easy", "understand", "escape-sequences-and-multiline-strings",
    "How does this appear on screen?\n\n"
    "```python\nprint(\"Line one\\nLine two\")\n```",
    "`\\n` is the escape sequence for a new line. Python does not print the characters `\\n`; it starts a fresh line there, so `Line one` and `Line two` appear on two separate lines.",
    ("Line one\\nLine two — the \\n shows as plain text",
     "Line one and Line two, on two separate lines",
     "LineoneLinetwo — the \\n is simply removed",
     "An error — \\n is not allowed inside print"),
    "B")

# Q9 - chaining strip().lower()
add(1, "hard", "analyze", "practical-text-processing",
    "What is printed? (Note the spaces around the text.)\n\n"
    "```python\nraw = \"  HELLO  \"\nclean = raw.strip().lower()\nprint(clean)\n```",
    "Methods can be chained. `strip()` first removes the spaces from both ends, giving `\"HELLO\"`, and `lower()` then makes it lowercase, giving `\"hello\"`. Each method passes its result to the next.",
    ("hello — strip removes the end spaces, then lower makes it lowercase",
     "  hello   — the spaces stay",
     "HELLO — lower does nothing here",
     "An error — you cannot chain two methods"),
    "A")

# Q10 - slicing: stop excluded + reverse
add(1, "hard", "analyze", "indexing-and-slicing",
    "What two lines does this print?\n\n"
    "```python\nword = \"Python\"\nprint(word[0:3])\nprint(word[::-1])\n```",
    "A slice `[0:3]` takes indexes 0, 1, 2 and stops just before 3, giving `Pyt`. The slice `[::-1]` steps backward through the whole string, reversing it to `nohtyP`.",
    ("Pyth then Python",
     "Pyt then Python",
     "Pyt then nohtyP — stop 3 is excluded, and [::-1] reverses",
     "Pyt on then nohtyP"),
    "C")


# =====================  SET 2  (Q11-Q40)  =====================

# Q11 - negative-index slice for last four
add(2, "medium", "analyze", "indexing-and-slicing",
    "Meera wants the last four digits of a phone number:\n\n"
    "```python\nphone = \"9876543210\"\nprint(phone[-4:])\n```\n\n"
    "What does this print?",
    "`[-4:]` means \"start four characters from the end and go to the end\". For `9876543210` that is the last four digits, `3210` — no matter how long the number is.",
    ("9876 — the first four digits",
     "3210 — the last four digits",
     "43210 — the last five digits",
     "An error — you cannot use a negative start"),
    "B")

# Q12 - immutability: build a new string
add(2, "medium", "analyze", "string-immutability",
    "What is printed?\n\n"
    "```python\nword = \"Python\"\nword = \"J\" + word[1:]\nprint(word)\n```",
    "`word[1:]` slices out everything from index 1 onward, `\"ython\"`. Joining `\"J\"` in front builds a brand new string `\"Jython\"`, which is stored back in `word`. The original was replaced, not edited.",
    ("Python — the original cannot change",
     "JPython",
     "An error — strings cannot be joined",
     "Jython — a new string is built from J plus ython"),
    "D")

# Q13 - replace swaps every occurrence
add(2, "medium", "analyze", "common-string-methods",
    "What does this print?\n\n"
    "```python\nsentence = \"tea, tea, and more tea\"\nprint(sentence.replace(\"tea\", \"coffee\"))\n```",
    "`replace` swaps *every* occurrence of the text, not just the first. All three copies of `tea` become `coffee`, giving `coffee, coffee, and more coffee`.",
    ("coffee, coffee, and more coffee — replace swaps every occurrence",
     "coffee, tea, and more tea — only the first is swapped",
     "tea, tea, and more coffee — only the last is swapped",
     "An error — replace takes only one argument"),
    "A")

# Q14 - isdigit on mixed text
add(2, "medium", "analyze", "common-string-methods",
    "What does this return?\n\n"
    "```python\nprint(\"12a\".isdigit())\n```",
    "`isdigit()` returns `True` only if *every* character is a digit. `\"12a\"` contains the letter `a`, so not all characters are digits, and it returns `False`.",
    ("True — it contains some digits",
     "12 — it returns just the digits",
     "False — not every character is a digit",
     "An error — isdigit needs a number"),
    "C")

# Q15 - split() default + len for word count
add(2, "medium", "analyze", "splitting-and-joining",
    "What number is printed?\n\n"
    "```python\nsentence = \"Python is really fun\"\nprint(len(sentence.split()))\n```",
    "`split()` with no argument cuts on spaces, turning the sentence into the list `['Python', 'is', 'really', 'fun']`. `len()` then counts the four items, so it prints 4.",
    ("1 — it counts the whole sentence as one",
     "18 — it counts the characters",
     "3 — it counts the spaces",
     "4 — split makes a list of words, and len counts them"),
    "D")

# Q16 - join with a separator
add(2, "medium", "analyze", "splitting-and-joining",
    "What does this print?\n\n"
    "```python\nwords = [\"Python\", \"is\", \"fun\"]\nprint(\"-\".join(words))\n```",
    "`\"-\".join(words)` glues the pieces together with a dash between each one, producing `Python-is-fun`. The string you call `join` on is the separator placed between items.",
    ("Python is fun — joined with spaces",
     "Python-is-fun — joined with a dash between each word",
     "['Python', 'is', 'fun'] — the list is unchanged",
     "An error — join needs a comma"),
    "B")

# Q17 - in operator
add(2, "easy", "understand", "searching-within-strings",
    "What does this print?\n\n"
    "```python\nemail = \"asha@email.com\"\nprint(\"@\" in email)\n```",
    "The `in` operator answers yes or no: is the text anywhere inside the string? The `@` is present, so it returns `True`.",
    ("True — the @ is somewhere in the text",
     "False — in only checks the very start",
     "4 — the position of the @",
     "An error — in cannot search a string"),
    "A")

# Q18 - count is case-sensitive
add(2, "hard", "analyze", "searching-within-strings",
    "How many does this count, and why?\n\n"
    "```python\ntext = \"Sea shells sold\"\nprint(text.count(\"s\"))\n```",
    "`count` searches for lowercase `s` and is case-sensitive, so the capital `S` in `Sea` does not count. The lowercase `s`s are two in `shells` and one in `sold` — three in total.",
    ("4 — it counts S and s together",
     "1 — only the first s",
     "3 — count is case-sensitive, so the capital S is not counted",
     "0 — s does not appear"),
    "C")

# Q19 - endswith
add(2, "easy", "understand", "searching-within-strings",
    "What does this print?\n\n"
    "```python\nfilename = \"report.pdf\"\nprint(filename.endswith(\".pdf\"))\n```",
    "`endswith` checks whether the string finishes with the given text and returns a boolean. `report.pdf` does end with `.pdf`, so it returns `True`.",
    (".pdf — it returns the ending",
     "An error — endswith needs an index",
     "False — endswith checks the start",
     "True — the filename ends with .pdf"),
    "D")

# Q20 - f-string alignment meaning
add(2, "medium", "understand", "string-formatting-and-fstrings",
    "In an f-string, a price list uses `{name:<10}` and `{price:>8}` to make tidy columns.\n\n"
    "What do the `<` and `>` mean here?",
    "After the colon, a number sets a minimum width and `<`, `>`, `^` set the alignment. `<` left-aligns the value in its width and `>` right-aligns it, which lines names up on the left and prices on the right.",
    ("> means round up, < means round down",
     "> right-aligns the value, < left-aligns it, within the given width",
     "> shows more decimals, < shows fewer",
     "They compare two numbers and give True or False"),
    "B")

# Q21 - triple-quoted multiline string
add(2, "medium", "understand", "escape-sequences-and-multiline-strings",
    "You need to store an email that has several lines and a blank line, exactly as typed, without scattering `\\n` after every sentence.\n\n"
    "Which tool fits best?",
    "A triple-quoted string (`\"\"\"...\"\"\"`) preserves everything between the triple quotes, including line breaks and blank lines, so the layout you type is the layout you get. That is exactly what multi-line text needs.",
    ("A triple-quoted string, which keeps every line break exactly as typed",
     "A single-quoted string with no line breaks",
     "The .split() method",
     "The len() function"),
    "A")

# Q22 - split then unpack
add(2, "hard", "analyze", "practical-text-processing",
    "What is printed?\n\n"
    "```python\nrecord = \"Asha,20,Pune\"\nname, age, city = record.split(\",\")\nprint(city)\n```",
    "`split(\",\")` gives three pieces: `'Asha'`, `'20'`, `'Pune'`. Those are unpacked in order into `name`, `age`, and `city`, so `city` holds `'Pune'`, which is printed.",
    ("Asha",
     "20",
     "Pune — split gives three pieces, unpacked into name, age, city",
     "An error — you cannot unpack a split"),
    "C")

# Q23 - IndexError out of range
add(2, "medium", "analyze", "indexing-and-slicing",
    "What happens here?\n\n"
    "```python\nword = \"cat\"\nprint(word[5])\n```",
    "`\"cat\"` has three characters, at indexes 0, 1, and 2. There is no index 5, so Python stops with an `IndexError` saying the index is out of range.",
    ("It prints t — the last letter",
     "It prints an empty string",
     "It prints 5",
     "It raises an IndexError — position 5 does not exist"),
    "D")

# Q24 - choosing quotes around an apostrophe
add(2, "medium", "apply", "what-is-a-string",
    "You need a string that contains an apostrophe: It's a sunny day.\n\n"
    "Which is the safe way to write it?",
    "Wrap the text in the quote style it does not contain. The text has an apostrophe (a single quote), so wrapping it in double quotes keeps the apostrophe from ending the string early: `\"It's a sunny day\"`.",
    ("'It's a sunny day' — single quotes around it",
     "\"It's a sunny day\" — wrap in double quotes so the apostrophe is safe",
     "It's a sunny day — no quotes are needed",
     "You cannot store an apostrophe in a string"),
    "B")

# Q25 - empty string length
add(2, "easy", "understand", "what-is-a-string",
    "What does this print?\n\n"
    "```python\nblank = \"\"\nprint(len(blank))\n```",
    "An empty string is two quotes with nothing between them, so it has no characters. `len(blank)` is therefore 0. (An empty string is also falsy.)",
    ("0 — an empty string has no characters",
     "1 — the quotes count as one character",
     "2 — the two quote marks",
     "An error — you cannot measure an empty string"),
    "A")

# Q26 - slice from start / to end
add(2, "medium", "analyze", "indexing-and-slicing",
    "What does this print?\n\n"
    "```python\nword = \"Python\"\nprint(word[:2], word[4:])\n```",
    "`word[:2]` means \"from the start up to but not including index 2\", giving `Py`. `word[4:]` means \"from index 4 to the end\", giving `on`. So it prints `Py on`.",
    ("Pyt then hon",
     "Py then hon",
     "Py then on — [:2] is the first two, [4:] is from index 4 to the end",
     "yt then on"),
    "C")

# Q27 - case-insensitive search
add(2, "hard", "apply", "searching-within-strings",
    "You want to check whether the word \"Fox\" appears in a sentence, no matter how it is capitalised. Remember that searching is case-sensitive.\n\n"
    "Which line does this correctly?",
    "Because searching is case-sensitive, you make both sides the same case first. Lower-casing the sentence and searching for the lowercase word — `\"fox\" in sentence.lower()` — matches `Fox`, `FOX`, or `fox` alike.",
    ("\"Fox\" in sentence",
     "sentence.count(\"Fox\")",
     "sentence.upper() in \"Fox\"",
     "\"fox\" in sentence.lower()"),
    "D")

# Q28 - strip removes only the ends
add(2, "hard", "analyze", "common-string-methods",
    "What is printed? (Shown between | and | to make the spaces clear.)\n\n"
    "```python\nmessy = \"  a b  \"\nprint(\"|\" + messy.strip() + \"|\")\n```",
    "`strip()` removes whitespace only from the two ends, not from the middle. The leading and trailing spaces go, but the single space between `a` and `b` stays, giving `|a b|`.",
    ("|ab| — all spaces are removed",
     "|a b| — only the end spaces are removed; the middle space stays",
     "|  a b  | — nothing is removed",
     "|a  b| — only the middle changes"),
    "B")

# Q29 - title()
add(2, "medium", "analyze", "common-string-methods",
    "What does this print?\n\n"
    "```python\nprint(\"hello world\".title())\n```",
    "`title()` capitalises the first letter of each word, so `\"hello world\"` becomes `\"Hello World\"`. (`upper()` would make it all capitals; `capitalize()` would raise only the very first letter.)",
    ("Hello World — title capitalises the first letter of each word",
     "HELLO WORLD — all capitals",
     "Hello world — only the first word",
     "hello world — unchanged"),
    "A")

# Q30 - f-string thousands separator
add(2, "medium", "apply", "string-formatting-and-fstrings",
    "What does this print?\n\n"
    "```python\nn = 1500000\nprint(f\"{n:,}\")\n```",
    "The `:,` format adds commas as thousands separators, making large numbers easier to read. 1500000 becomes `1,500,000`.",
    ("1500000 — unchanged",
     "1.5 — shortened",
     "1,500,000 — the comma groups the thousands",
     "An error — you cannot format an integer"),
    "C")

# Q31 - char count without spaces
add(2, "hard", "analyze", "practical-text-processing",
    "What number is printed?\n\n"
    "```python\ntext = \"a b c\"\nprint(len(text.replace(\" \", \"\")))\n```",
    "`replace(\" \", \"\")` removes every space, turning `\"a b c\"` into `\"abc\"`. `len(\"abc\")` is 3. This is the common way to count characters while ignoring spaces.",
    ("5 — it counts the spaces too",
     "3 — the spaces are removed first, leaving abc",
     "2 — it counts the spaces",
     "An error — replace needs two words"),
    "B")

# Q32 - immutability concept via id / reassign
add(2, "medium", "understand", "string-immutability",
    "After `word = \"J\" + word[1:]`, the variable `word` points to a brand new string and `id(word)` has changed.\n\n"
    "What does this tell you about strings?",
    "A changed `id` means the variable was pointed at a new object, not that the old string was edited. Strings are immutable: you build a new one and repoint the variable, while the original stays untouched.",
    ("The old string was edited in place",
     "Strings can be changed with slicing",
     "id() always stays the same for strings",
     "Strings are immutable — you build a new one and repoint the variable, never editing the old"),
    "D")

# Q33 - split(",") vs split()
add(2, "medium", "understand", "splitting-and-joining",
    "What is the difference between `text.split(\",\")` and `text.split()` with no argument?",
    "`split(\",\")` cuts the string wherever it finds a comma. `split()` with no argument cuts on spaces instead, which is the easy way to break a sentence into words.",
    ("split(\",\") cuts at each comma; split() with nothing cuts at spaces",
     "They do exactly the same thing",
     "split() cuts at commas; split(\",\") cuts at spaces",
     "split() reverses the string"),
    "A")

# Q34 - escape \t and \"
add(2, "medium", "analyze", "escape-sequences-and-multiline-strings",
    "What appears on screen?\n\n"
    "```python\nprint(\"Name:\\tAsha\")\nprint(\"She said \\\"hi\\\"\")\n```",
    "`\\t` inserts a tab, so a gap appears between `Name:` and `Asha`. `\\\"` puts a real double quote inside the double-quoted string, so the second line shows `She said \"hi\"` with the quote marks visible.",
    ("Name:\\tAsha then She said \\\"hi\\\" — the codes show as plain text",
     "NameAsha then She said hi — the codes are deleted",
     "Name: then a tab then Asha, and She said \"hi\" with real quote marks",
     "An error — you cannot mix \\t and \\\""),
    "C")

# Q35 - raw string
add(2, "hard", "analyze", "escape-sequences-and-multiline-strings",
    "What does this print, and why?\n\n"
    "```python\npath = r\"C:\\new\\table\"\nprint(path)\n```",
    "The `r` prefix makes a raw string, which turns off escape sequences. So `\\n` and `\\t` are not treated as newline and tab — every backslash is kept literally, printing `C:\\new\\table`.",
    ("C: then a new line, then ew, a tab, then able — \\n and \\t act as codes",
     "An error — backslashes are not allowed in strings",
     "C:newtable — the backslashes are removed",
     "C:\\new\\table — the r prefix turns off escaping, so the backslashes stay literal"),
    "D")

# Q36 - find returns a position
add(2, "medium", "analyze", "searching-within-strings",
    "What is printed?\n\n"
    "```python\nemail = \"asha@email.com\"\nprint(email.find(\"@\"))\n```",
    "`find` returns the index where the text first appears. Counting from 0: a(0), s(1), h(2), a(3), @(4). So the `@` is at index 4 and `find` returns 4.",
    ("True — @ is present",
     "4 — find returns the index where @ first appears",
     "1 — it counts one @",
     "-1 — @ is not found"),
    "B")

# Q37 - startswith
add(2, "easy", "understand", "searching-within-strings",
    "What does this print?\n\n"
    "```python\nname = \"Meera\"\nprint(name.startswith(\"Me\"))\n```",
    "`startswith` checks whether the string begins with the given text and returns a boolean. `Meera` does begin with `Me`, so it returns `True`.",
    ("True — the name begins with Me",
     "False — startswith checks the end",
     "Me — it returns the matched part",
     "0 — the starting index"),
    "A")

# Q38 - f-string percentage
add(2, "hard", "apply", "string-formatting-and-fstrings",
    "What does this print?\n\n"
    "```python\nratio = 0.8734\nprint(f\"{ratio:.1%}\")\n```",
    "The `.1%` format multiplies by 100, adds a percent sign, and shows one decimal place. 0.8734 becomes 87.34%, rounded to one decimal: `87.3%`.",
    ("0.9%",
     "0.8734%",
     "87.3% — .1% shows a percentage with one decimal place",
     "8.7%"),
    "C")

# Q39 - chaining clean-up, inner spaces stay
add(2, "medium", "analyze", "practical-text-processing",
    "What is printed?\n\n"
    "```python\nraw = \"  Hello,  World!  \"\nprint(raw.strip().lower().replace(\",\", \"\"))\n```",
    "Each method feeds the next: `strip()` removes the outer spaces, `lower()` lowercases everything, and `replace(\",\", \"\")` drops the comma. The two spaces between the words are in the middle, so they stay: `hello  world!`.",
    ("Hello, World! — nothing changes",
     "helloworld! — every space and the comma are removed",
     "hello, world! — the comma stays",
     "hello  world! — trimmed, lowercased, and the comma removed (inner spaces stay)"),
    "D")

# Q40 - a string is an ordered sequence
add(2, "medium", "understand", "what-is-a-string",
    "Why can a for loop step through a string one character at a time, as in `for letter in \"cat\"`?",
    "A string is an ordered sequence of characters, each sitting in its own position one after another. That order is what lets a loop hand you each character in turn, just as it would each item in a list.",
    ("Because strings are secretly numbers",
     "Because a string is an ordered sequence of characters, each in its own position",
     "Because Python converts the string to a list first",
     "Because len() splits it into letters"),
    "B")


# =====================  Build workbook  =====================

assert len(Q) == 40, f"expected 40 questions, got {len(Q)}"

os.makedirs(OUT_DIR, exist_ok=True)
shutil.copyfile(TEMPLATE, OUT_FILE)

wb = openpyxl.load_workbook(OUT_FILE)
ws = wb.active
ws.title = SHEET_NAME
ws.delete_rows(2, ws.max_row)  # keep header, drop sample rows

set_counter = {1: 0, 2: 0}
letters = []
answer_tally = Counter()

for setn, diff, bloom, sub, desc, expl, opts, ans in Q:
    set_counter[setn] += 1
    title = f"Python - MCQ - 5.{setn}.{set_counter[setn]}"
    tag = f"python - Set {setn}"
    answer_tally[ans] += 1
    letters.append(ans)
    ws.append([
        title, desc, expl, SCORE[diff], "published", diff, bloom,
        tag, "python", "strings", sub, None,
        opts[0], opts[1], opts[2], opts[3], L[ans],
    ])

wb.save(OUT_FILE)

# =====================  Validation  =====================
print("Saved:", OUT_FILE)
print("Total questions:", len(Q))
print("Answer distribution:", dict(sorted(answer_tally.items())))

run = max_run = 1
for a, b in zip(letters, letters[1:]):
    run = run + 1 if a == b else 1
    max_run = max(max_run, run)
print("Max consecutive same letter:", max_run)

assert all(answer_tally[x] == 10 for x in "ABCD"), answer_tally
assert max_run <= 2, "answer letter repeats more than twice consecutively"

chk = openpyxl.load_workbook(OUT_FILE)
cs = chk[SHEET_NAME]
rows = list(cs.iter_rows(values_only=True))
assert rows[0] == ('title', 'description', 'explanation', 'score', 'status',
                   'difficulty', 'bloomTaxonomy', 'tags', 'subjects', 'topics',
                   'subTopics', 'companies', 'option1', 'option2', 'option3',
                   'option4', 'answer'), rows[0]
assert len(rows) == 41, len(rows)
for r in rows[1:]:
    assert len(set(r[12:16])) == 4, f"duplicate options in {r[0]}"
    assert r[12:16][r[16] - 1], f"empty correct option in {r[0]}"
set1 = [r[10] for r in rows[1:] if r[7] == 'python - Set 1']
allsubs = set(r[10] for r in rows[1:])
print("Difficulty mix:", dict(Counter(r[5] for r in rows[1:])))
print("Set 1:", sum(1 for r in rows[1:] if r[7] == 'python - Set 1'),
      "| Set 2:", sum(1 for r in rows[1:] if r[7] == 'python - Set 2'))
print("Set 1 lesson coverage:", len(set(set1)), "/", len(allsubs),
      "| missing:", allsubs - set(set1) or "none")
print("Validation passed.")
