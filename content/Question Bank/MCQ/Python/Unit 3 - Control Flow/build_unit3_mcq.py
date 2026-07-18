"""Build the Unit 3 - Control Flow MCQ workbook from the shared template.

Grounded strictly in content/Semester 1/Unit 3 - Control Flow reading materials:
if / if-else / elif ordering / nested & guard clauses / match-case /
truthiness / ternary expressions / input validation with .isdigit().

Writing rules applied to every question:
* Simple English, short sentences (first-year reading level).
* The rule is shown through an example or outcome where that forces thinking,
  instead of naming the operator in words (which would let a student guess).
* Trace-values sit on or near boundaries, so a wrong reasoning path gives a
  different answer.
* Each distractor maps to one specific, nameable misconception.
* Every question demands its own distinct reasoning step.

Set 1 (Q1-10)  -> tag "python - Set 1", titles 3.1.1 .. 3.1.10
Set 2 (Q11-40) -> tag "python - Set 2", titles 3.2.1 .. 3.2.30
Answer field is 1..4 (A=1, B=2, C=3, D=4).
"""

import os
import shutil
from collections import Counter

import openpyxl

TEMPLATE = "content/Question Bank/Template/questions-mcq-template.xlsx"
OUT_DIR = "content/Question Bank/MCQ/Python/Unit 3 - Control Flow"
OUT_FILE = os.path.join(OUT_DIR, "Unit 3 - Control Flow - MCQ.xlsx")
SHEET_NAME = "Python - MCQ - 3"

L = {"A": 1, "B": 2, "C": 3, "D": 4}
SCORE = {"easy": 5, "medium": 8, "hard": 10}

Q = []


def add(setn, diff, bloom, sub, desc, expl, opts, ans):
    Q.append((setn, diff, bloom, sub, desc, expl, opts, ans))


# =====================  SET 1  (Q1-Q10)  =====================

# Q1 - truthy/falsy of the text "0"
add(1, "medium", "understand", "truthiness-and-boolean-logic",
    "A tracking page starts a lookup only if the customer typed an ID. Here the ID is the text \"0\":\n\n"
    "```python\ntracking_id = \"0\"\nif tracking_id:\n    print(\"Tracking parcel\")\nelse:\n    print(\"No ID entered\")\n```\n\n"
    "A teammate says the lookup will be skipped, because \"0 means empty\". What does the code print?",
    "The value is the text `\"0\"`, which has one character, so it is not an empty string. A non-empty string is truthy, so the `if` runs and prints `Tracking parcel`. Only the number `0` is falsy, not the text `\"0\"` — so the teammate is wrong.",
    ("No ID entered — the text \"0\" counts as empty",
     "An error — you cannot put a string after if",
     "Tracking parcel — \"0\" has one character, so it is not empty",
     "No ID entered — Python turns \"0\" into the number 0 first"),
    "C")

# Q2 - completing a condition; rule shown by outcome, boundary value in code
add(1, "easy", "apply", "the-if-statement",
    "An online store gives free shipping once a cart goes past ₹999. So a cart of ₹999 must still pay for shipping, but a cart of ₹1000 gets it free. The cart below is exactly ₹999:\n\n"
    "```python\ncart_total = 999\nif ________:\n    print(\"Free shipping unlocked\")\nelse:\n    print(\"Pay for shipping\")\n```\n\n"
    "Which condition gives the right result for this ₹999 cart?",
    "A ₹999 cart must NOT get free shipping. `cart_total > 999` is `999 > 999` → False, so it correctly prints `Pay for shipping`, and a ₹1000 cart would pass. `>= 999` would wrongly free the ₹999 cart, `< 999` reverses the rule, and `== 999` matches only that one value.",
    ("cart_total > 999",
     "cart_total >= 999",
     "cart_total < 999",
     "cart_total == 999"),
    "A")

# Q3 - elif ordering: the first true condition wins
add(1, "medium", "analyze", "elif-multi-way-branching",
    "A grading chain should label 90 and above as \"A\", 75-89 as \"B\", and 60-74 as \"C\". A developer writes the checks from the lowest boundary upward, and a top scorer of 95 comes out labelled \"C\":\n\n"
    "```python\nmarks = 95\nif marks >= 60:\n    print(\"C\")\nelif marks >= 75:\n    print(\"B\")\nelif marks >= 90:\n    print(\"A\")\n```\n\n"
    "Why does 95 print \"C\"?",
    "An elif chain is checked top to bottom, and the first true condition wins — every branch after it is skipped. For 95, `marks >= 60` is already true, so it prints \"C\" and never reaches the checks for \"B\" or \"A\". The fix is to order the conditions from the highest threshold down.",
    ("Because 95 is not greater than 90",
     "Because the first true condition wins — marks >= 60 matches first, so the rest are skipped",
     "Because an elif chain prints every branch that is true",
     "Because the else branch is missing"),
    "B")

# Q4 - spot the boundary bug (from a complaint)
add(1, "medium", "analyze", "if-else",
    "A shop gives a discount on any bill of ₹2000 or more. But customers billed exactly ₹2000 keep complaining that they never get it:\n\n"
    "```python\nbill = 2000\nif bill > 2000:\n    print(\"Discount applied\")\nelse:\n    print(\"No discount\")\n```\n\n"
    "What is the bug, and what shows for a ₹2000 bill?",
    "\"₹2000 or more\" must include 2000, so the test needs `>=`. With `>`, a bill of exactly 2000 fails and the `else` runs, showing `No discount` — the exact complaint. Changing `>` to `>=` fixes it.",
    ("> should be !=; it shows No discount",
     "There is no bug; it shows Discount applied",
     "> should be >=; it shows No discount",
     "> should be ==; it shows Discount applied"),
    "C")

# Q5 - unexpected behaviour from independent ifs
add(1, "hard", "analyze", "if-else",
    "A fitness band should give one badge per day. But users with high step counts say several badges pop up at once. Here is the badge code for a 12,000-step day:\n\n"
    "```python\nsteps = 12000\nif steps >= 5000:\n    print(\"Bronze\")\nif steps >= 8000:\n    print(\"Silver\")\nif steps >= 10000:\n    print(\"Gold\")\n```\n\n"
    "What does it print, and why do users see a pile-up?",
    "These are three separate `if` statements, not one `elif` chain. At 12,000 steps every condition is True, so all three run and print `Bronze`, `Silver`, and `Gold` — the pile-up. Using `elif` for the last two would let only one badge show.",
    ("Gold — only the top badge shows",
     "Bronze — the first true line stops the rest",
     "Nothing — the ranges clash",
     "Bronze, Silver, and Gold — each if runs on its own"),
    "D")

# Q6 - selecting the most appropriate structure
add(1, "medium", "understand", "match-case",
    "An app must turn one typed word — \"lights\", \"fan\", \"tv\", or \"ac\" — into an action, and reply \"Not recognised\" for anything else. Every choice compares the same word against a fixed value.\n\n"
    "Which structure fits this best?",
    "Comparing one value to several fixed options, with a default for the rest, is exactly what `match-case` is for — `case _` handles anything unknown. Four separate `if`s could fire more than once and have no clean default, one `if` with `or` cannot pick a different action per word, and deep nesting only adds clutter.",
    ("A match-case block with a case _ for anything else",
     "Four separate if statements, one per word",
     "One if that joins all four words with or",
     "An if-else nested four levels deep"),
    "A")

# Q7 - operator precedence; data chosen to expose the missing brackets
add(1, "hard", "analyze", "truthiness-and-boolean-logic",
    "An insurer approves a claim only if the policy is active AND the claim is \"medical\" or \"accident\". A coder thinks all four lines below mean the same thing. To settle it, they test an inactive policy with an \"accident\" claim — which must be rejected:\n\n"
    "```python\nactive = False\nclaim_type = \"accident\"\nif ________:\n    print(\"Claim approved\")\n```\n\n"
    "Which condition correctly rejects this claim?",
    "`and` is checked before `or`. Line A reads as `(active and medical) or accident` → `(False and ...) or True` → True, so it wrongly approves. Line B puts brackets around the two claim types: `False and (False or True)` → `False and True` → False, so it rejects — correct. Line C's leading `or` drops the active rule, and line D needs both claim types at once, which is impossible.",
    ("active and claim_type == \"medical\" or claim_type == \"accident\"",
     "active and (claim_type == \"medical\" or claim_type == \"accident\")",
     "active or claim_type == \"medical\" and claim_type == \"accident\"",
     "active and claim_type == \"medical\" and claim_type == \"accident\""),
    "B")

# Q8 - equivalence: nested if (no inner else) vs and
add(1, "medium", "analyze", "nested-conditions-and-guard-clauses",
    "A garden system waters a bed only when the soil is dry AND the tank has water. Two coders write it differently. Notice the inner `if` has no `else`:\n\n"
    "```python\n# Version A\nif soil_dry:\n    if tank_has_water:\n        print(\"Watering\")\n\n# Version B\nif soil_dry and tank_has_water:\n    print(\"Watering\")\n```\n\n"
    "Do the two versions always behave the same?",
    "A nested `if` with no inner `else` is the same as one `if` joined by `and`: both water only when soil is dry and the tank has water, and both do nothing otherwise. So they always match; Version B is just easier to read.",
    ("No — Version A also waters when the tank is empty",
     "No — Version B skips the soil check",
     "No — Version A checks the tank even when the soil is wet",
     "Yes — both water only when the soil is dry and the tank has water"),
    "D")

# Q9 - comparing two implementations with a tie-breaker input
add(1, "hard", "analyze", "input-validation",
    "A site accepts a coupon only if it is 8 characters long and not already used. Two coders write the check, then test a 5-character coupon that has never been used:\n\n"
    "```python\nused = False\ncode = input(\"Enter coupon: \")\n\n# Option 1\nif len(code) != 8:\n    print(\"Wrong length\")\nelif used:\n    print(\"Already used\")\nelse:\n    print(\"Coupon applied\")\n\n# Option 2\nif len(code) != 8 and used:\n    print(\"Invalid coupon\")\nelse:\n    print(\"Coupon applied\")\n```\n\n"
    "Which version rejects the bad coupon, and why does the other let it through?",
    "Option 1 checks the two rules one after another, so the wrong length is caught. Option 2 joins them with `and`, so it only rejects a coupon that is BOTH the wrong length AND already used. A 5-character unused coupon fails only one rule, so Option 2 prints `Coupon applied` and lets it through.",
    ("Option 1 rejects it; Option 2 only rejects a coupon that is the wrong length AND already used",
     "Both reject it; the two versions behave the same",
     "Option 2 rejects it; Option 1 just adds extra branches for no reason",
     "Option 1 rejects it; Option 2 fails because only one of the two checks matters"),
    "A")

# Q10 - selecting the one valid / smallest correct line
add(1, "easy", "apply", "conditional-ternary-expressions",
    "A toll booth needs a one-line label: \"Heavy\" above 3000 kg, otherwise \"Light\". Three of the four lines below are broken — one uses statement syntax, one swaps the labels, one is missing a part:\n\n"
    "```python\nweight = 2500\ncategory = ________\nprint(category)\n```\n\n"
    "Which line is the correct one?",
    "A conditional expression is written `value_if_true if condition else value_if_false`, so `\"Heavy\" if weight > 3000 else \"Light\"` is right. Line A is `if`/`else` statement syntax, not an expression; line B swaps the two labels; line D is missing the `else` part.",
    ("if weight > 3000: \"Heavy\" else: \"Light\"",
     "\"Light\" if weight > 3000 else \"Heavy\"",
     "\"Heavy\" if weight > 3000 else \"Light\"",
     "\"Heavy\" if weight > 3000"),
    "C")


# =====================  SET 2  (Q11-Q40)  =====================

# Q11 - choosing the input that exposes >= vs >
add(2, "medium", "analyze", "if-else",
    "A ward monitor should alert when the temperature reaches 38.0°C. A tester wants the single reading that would reveal the bug if someone wrote `>` instead of `>=`:\n\n"
    "```python\ndef needs_alert(temp):\n    return temp >= 38.0\n```\n\n"
    "Which reading must be in the test?",
    "`>=` and `>` give different answers only at the exact number. `needs_alert(38.0)` is True with `>=`, but would be False with `>`. Any reading away from 38.0 gives the same answer either way, so only 38.0 catches the mistake.",
    ("36.5",
     "39.2",
     "41.0",
     "38.0"),
    "D")

# Q12 - the smallest correct elif (nudged to medium)
add(2, "medium", "apply", "elif-multi-way-branching",
    "A courier charges by weight: under 1 kg is \"Light\", 1 to 5 kg is \"Standard\", over 5 kg is \"Heavy\". The `if` above already handled `weight < 1`, so anything reaching the `elif` is at least 1 kg:\n\n"
    "```python\nweight = 3.0\nif weight < 1:\n    print(\"Light\")\nelif ________:\n    print(\"Standard\")\nelse:\n    print(\"Heavy\")\n```\n\n"
    "Which is the shortest condition that still works?",
    "Since `weight < 1` is already ruled out, the value is guaranteed to be 1 or more — no need to re-check the lower end. Only the top edge is left, and 5 is included, so `weight <= 5` is enough. `weight < 5` wrongly pushes 5 into \"Heavy\"; the others re-test a lower bound that is already certain.",
    ("weight > 1 and weight < 5",
     "weight <= 5",
     "weight >= 1 and weight <= 5",
     "weight < 5"),
    "B")

# Q13 - choosing input that exposes a str-vs-int bug
add(2, "hard", "analyze", "input-validation",
    "During testing, this age check is always called with whole numbers, and every call passes. But in the real app the value comes straight from `input()`:\n\n"
    "```python\ndef check_can_vote(age):\n    return \"Eligible\" if age >= 18 else \"Not eligible\"\n```\n\n"
    "Which single call behaves like the real app and reveals the hidden problem?",
    "`input()` always returns a string, so the real app passes text like `\"25\"`. Comparing `\"25\" >= 18` compares a string to a number, which raises a `TypeError` in Python 3. Calls with plain numbers — normal or extreme — never hit this, which is why the tests missed it.",
    ("check_can_vote(\"25\")",
     "check_can_vote(25)",
     "check_can_vote(16)",
     "check_can_vote(0)"),
    "A")

# Q14 - spot the reversed comparison
add(2, "easy", "analyze", "if-else",
    "A library should charge a late fee when a book is kept for more than 14 days. But a book returned after 20 days shows \"Returned on time\":\n\n"
    "```python\ndays = 20\nif days < 14:\n    print(\"Late fee applies\")\nelse:\n    print(\"Returned on time\")\n```\n\n"
    "What is wrong with the check?",
    "The comparison points the wrong way. `days < 14` is True only for short loans, so a 20-day loan takes the `else` and is called on time. The fee is for *more than* 14 days, so the test should be `days > 14`.",
    ("It should be >= 14 to include the 14th day",
     "Nothing is wrong — 20 days is within the limit",
     "< should be > — the fee is for more than 14 days, not fewer",
     "The else branch should print the fee"),
    "C")

# Q15 - and/or logic bug
add(2, "medium", "analyze", "truthiness-and-boolean-logic",
    "A weather station should flag any reading below -50 or above 150. But a reading of 200 was called \"normal\":\n\n"
    "```python\nreading = 200\nif reading < -50 and reading > 150:\n    print(\"Sensor fault\")\nelse:\n    print(\"Reading normal\")\n```\n\n"
    "Why did it not flag 200?",
    "One number can never be below -50 AND above 150 at the same time, so the `and` is always False and every reading is called normal. It should be `or`: flag when the reading is below -50 OR above 150.",
    ("The comparisons are backwards",
     "and should be or — no number is below -50 and above 150 at once",
     "The else needs its own condition",
     "Both < and > should become <= and >="),
    "B")

# Q16 - final value from a ternary with and (trap: ignoring the and)
add(2, "medium", "apply", "conditional-ternary-expressions",
    "A player shows 4K only for a premium user on a line faster than 25 Mbps. This user is premium, but the line is only 18 Mbps:\n\n"
    "```python\npremium = True\nbandwidth = 18\nquality = \"4K\" if premium and bandwidth > 25 else \"HD\"\nprint(quality)\n```\n\n"
    "What is quality set to?",
    "The condition is `premium and bandwidth > 25`. `premium` is True but `bandwidth > 25` is `18 > 25` → False, so `True and False` → False. The expression takes the `else` value, `\"HD\"`. Both parts of the `and` must be True to get 4K.",
    ("4K — being premium is enough",
     "An error — you cannot use and inside a ternary",
     "4K — the second part of the and is ignored",
     "HD — the line is not faster than 25, so the and is False"),
    "D")

# Q17 - analysing nested conditions
add(2, "medium", "understand", "nested-conditions-and-guard-clauses",
    "A lift moves only when the doors are shut, and even then it refuses if the load is over 800 kg. Right now the doors are shut and the load is 850 kg:\n\n"
    "```python\ndoors_closed = True\nweight = 850\n\nif doors_closed:\n    if weight <= 800:\n        print(\"Lift moving\")\n    else:\n        print(\"Overloaded\")\nelse:\n    print(\"Doors open\")\n```\n\n"
    "Which one message shows?",
    "The doors are shut, so the outer `if` runs. Inside, `weight <= 800` is `850 <= 800` → False, so the inner `else` runs and shows `Overloaded`. The outer `else` (\"Doors open\") belongs to the door check and is skipped.",
    ("Overloaded",
     "Lift moving",
     "Doors open",
     "Both Overloaded and Doors open"),
    "A")

# Q18 - completing a compound condition with not
add(2, "medium", "apply", "truthiness-and-boolean-logic",
    "A gym gate opens only for a member who is active and has not already entered — so one pass cannot be used twice. This member is active and has not entered yet:\n\n"
    "```python\nactive = True\nchecked_in = False\n\nif ________:\n    print(\"Access granted\")\nelse:\n    print(\"Access denied\")\n```\n\n"
    "Which condition opens the gate now but blocks a second entry?",
    "Both parts must be true: the member is active, and they have not entered (`not checked_in`). Here `active and not checked_in` → `True and not False` → `True and True` → True, so the gate opens. If they had already entered, `not checked_in` would be False and the gate would stay shut.",
    ("active or not checked_in",
     "active and checked_in",
     "active and not checked_in",
     "not active and not checked_in"),
    "C")

# Q19 - equivalence: ternary with swapped labels
add(2, "medium", "analyze", "conditional-ternary-expressions",
    "A coder rewrote a four-line decision as a single line, and a reviewer says the meaning flipped. The applicant's score is 720:\n\n"
    "```python\n# Version A\nif score >= 700:\n    decision = \"Approved\"\nelse:\n    decision = \"Rejected\"\n\n# Version B\ndecision = \"Rejected\" if score >= 700 else \"Approved\"\n```\n\n"
    "What does each version store for score = 720?",
    "In Version A, `720 >= 700` is True, so `decision = \"Approved\"`. In Version B the labels are the wrong way round: when the condition is True it returns `\"Rejected\"`. So the two disagree — the reviewer is right. A correct one-liner would read `\"Approved\" if score >= 700 else \"Rejected\"`.",
    ("Both store \"Approved\"",
     "Version A stores \"Approved\", Version B stores \"Rejected\"",
     "Both store \"Rejected\"",
     "Version A stores \"Rejected\", Version B stores \"Approved\""),
    "B")

# Q20 - selecting the smallest correct repair
add(2, "medium", "analyze", "elif-multi-way-branching",
    "A post office charges: up to 250 g → ₹20, 251 to 500 g → ₹40, over 500 g → ₹60. A 500 g parcel is being charged ₹60 instead of ₹40:\n\n"
    "```python\ngrams = 500\nif grams <= 250:\n    cost = 20\nelif grams < 500:\n    cost = 40\nelse:\n    cost = 60\nprint(cost)\n```\n\n"
    "Which single smallest change fixes the 500 g parcel?",
    "The middle band uses `grams < 500`, so exactly 500 fails it and drops to the `else`, costing ₹60. The ₹40 band is meant to include 500, so changing `< 500` to `<= 500` is the smallest fix. Reordering or touching the 250 line does not help, and turning the `else` into `grams > 500` would leave 500 matching no branch at all.",
    ("Put the else first",
     "Change grams <= 250 to grams < 250",
     "Change the else to elif grams > 500",
     "Change grams < 500 to grams <= 500"),
    "D")

# Q21 - selecting the structure that guarantees exactly one message
add(2, "medium", "understand", "elif-multi-way-branching",
    "A dashboard shows one label per rating: under 3 is \"Improve\", 3 to 4 is \"Meets\", over 4 is \"Exceeds\". It must never show two labels, and never show none.\n\n"
    "Which structure guarantees exactly one label for any rating?",
    "An `if`-`elif`-`else` chain runs only the first matching branch, and the `else` covers everything left over — so exactly one label always shows. Separate `if`s could fire more than once, one `if` with `or` cannot pick a different label per range, and `match-case` is for fixed values, not ranges.",
    ("An if-elif-else chain",
     "Three separate if statements",
     "One if with the ranges joined by or",
     "A match-case that compares the rating to each number"),
    "A")

# Q22 - condition order bug
add(2, "hard", "analyze", "elif-multi-way-branching",
    "A data plan charges: up to 2 GB is free, up to 10 GB is ₹100, over 10 GB is ₹300. A user on 1.5 GB — clearly free — was charged ₹100:\n\n"
    "```python\ngb = 1.5\nif gb <= 10:\n    charge = 100\nelif gb <= 2:\n    charge = 0\nelse:\n    charge = 300\nprint(charge)\n```\n\n"
    "Why was the free user charged?",
    "`gb <= 10` is True for anything up to 10 GB, including 1.5 GB, so the first branch runs and charges ₹100. The `elif gb <= 2` can never be reached for any value at or below 10. Range checks must go from smallest to largest: test `<= 2` first, then `<= 10`, then `else`.",
    ("The else should say gb > 10",
     "Every <= should be <",
     "The order is wrong — gb <= 10 catches small usage first",
     "The free and top tiers need a lower bound with and"),
    "C")

# Q23 - indentation / else attached to the wrong if
add(2, "hard", "analyze", "nested-conditions-and-guard-clauses",
    "A lock should say \"Invalid card\" when a valid card is used with the wrong PIN. Instead it stays completely silent:\n\n"
    "```python\ncard_valid = True\npin = \"0000\"\n\nif card_valid:\n    if pin == \"1234\":\n        print(\"Unlocked\")\nelse:\n    print(\"Invalid card\")\n```\n\n"
    "What does this print for a valid card and the wrong PIN?",
    "The `else` lines up with the outer `if card_valid`, not the PIN check. The card is valid, so that `else` is skipped. Inside, `pin == \"1234\"` is False, so nothing runs there either — hence the silence. A wrong PIN needs its own `else` under the inner `if`.",
    ("Invalid card",
     "Nothing",
     "Unlocked",
     "Both Unlocked and Invalid card"),
    "B")

# Q24 - match-case is case-sensitive
add(2, "medium", "understand", "match-case",
    "A kiosk matches typed language codes, which are all written in lowercase. A traveller types \"EN\":\n\n"
    "```python\nlang = \"EN\"\nmatch lang:\n    case \"en\":\n        print(\"English selected\")\n    case \"hi\":\n        print(\"Hindi selected\")\n    case \"ta\":\n        print(\"Tamil selected\")\n    case _:\n        print(\"Language not supported\")\n```\n\n"
    "What prints, and why?",
    "`match-case` compares for an exact match, and capital letters count. `\"EN\"` does not equal `\"en\"`, `\"hi\"`, or `\"ta\"`, so none of the named cases match and `case _` runs, printing `Language not supported`. Using `lang.lower()` first would fix it.",
    ("English selected — match ignores capitals",
     "Nothing — match errors when no case matches",
     "English selected — the _ case jumps to the first case",
     "Language not supported — \"EN\" does not equal \"en\", so no case matches"),
    "D")

# Q25 - truthiness of a spaces-only string
add(2, "medium", "understand", "truthiness-and-boolean-logic",
    "A form should block a blank subject. But a ticket goes through even when the subject is just two spaces:\n\n"
    "```python\nsubject = \"  \"\nif subject:\n    print(\"Ticket created\")\nelse:\n    print(\"Subject required\")\n```\n\n"
    "What happens, and why?",
    "For a string, only the empty string `\"\"` is falsy. `\"  \"` has two space characters, so it is not empty and counts as truthy — the ticket is created. To block spaces-only input, the developer should test `if subject.strip():`.",
    ("Ticket created — spaces still count as characters, so the string is not empty",
     "Subject required — a string of only spaces counts as empty",
     "An error — spaces cannot be tested in an if",
     "Subject required — Python removes the spaces first"),
    "A")

# Q26 - tracing multiple conditions in a chain
add(2, "hard", "analyze", "truthiness-and-boolean-logic",
    "A dispatcher checks a robot in order. Right now it has an order but no free slot yet:\n\n"
    "```python\nhas_order = True\nslot_free = False\nzone = \"\"\n\nif has_order and slot_free:\n    print(\"Robot dispatched\")\nelif has_order and not slot_free:\n    print(\"Order queued — waiting for a slot\")\nelif not has_order and zone:\n    print(\"Idle move to zone\")\nelse:\n    print(\"Robot idle\")\n```\n\n"
    "Which one line prints?",
    "First test: `has_order and slot_free` → `True and False` → False. Next: `has_order and not slot_free` → `True and True` → True, so it prints `Order queued — waiting for a slot`. Once a branch runs, the rest are skipped; the empty `zone` is never even checked.",
    ("Robot idle",
     "Order queued — waiting for a slot",
     "Idle move to zone",
     "Robot dispatched"),
    "B")

# Q27 - truthiness of the number 0
add(2, "easy", "understand", "truthiness-and-boolean-logic",
    "A phone hides its alert badge when nothing is unread. Right now the unread counter is 0:\n\n"
    "```python\nnotifications = 0\nif notifications:\n    print(\"You have new alerts\")\nelse:\n    print(\"No new alerts\")\n```\n\n"
    "What prints?",
    "The number `0` is falsy, so `if notifications:` is False and the `else` runs, printing `No new alerts`. Any other number would be truthy and show the alert.",
    ("You have new alerts — any number is truthy",
     "An error — a number is not a boolean",
     "No new alerts — 0 is falsy",
     "You have new alerts — 0 still counts as something"),
    "C")

# Q28 - completing an inclusive-range condition (rule shown by which values count)
add(2, "medium", "apply", "the-if-statement",
    "A clinic calls a reading \"Normal\" only from 90 to 120, and both 90 and 120 themselves count as normal:\n\n"
    "```python\nsystolic = 110\nif ________:\n    print(\"Normal\")\nelse:\n    print(\"Check with doctor\")\n```\n\n"
    "Which condition keeps 90 and 120 as normal, without calling every value normal?",
    "The band includes both ends, and Python lets you chain it as `90 <= systolic <= 120`. Line A uses strict `<`/`>` and wrongly drops 90 and 120; line B with `or` is true for every number; line C can only be true if one value equals both 90 and 120, which is impossible.",
    ("systolic > 90 and systolic < 120",
     "systolic >= 90 or systolic <= 120",
     "systolic == 90 and systolic == 120",
     "90 <= systolic <= 120"),
    "D")

# Q29 - identifying an unreachable branch
add(2, "medium", "analyze", "elif-multi-way-branching",
    "A colleague says one line in this shipping code can never run, whatever the tier is:\n\n"
    "```python\ntier = \"gold\"\nif tier == \"gold\":\n    print(\"Free next-day delivery\")\nelif tier == \"gold\":\n    print(\"Free standard delivery\")\nelse:\n    print(\"Paid delivery\")\n```\n\n"
    "Which line can never run, and what does a \"gold\" customer see?",
    "The `elif` repeats the exact test of the `if`. If the `if` is True, the `elif` is never reached; if the `if` is False, the same `elif` is also False. So the second line can never run. For `\"gold\"`, the `if` runs and shows `Free next-day delivery`.",
    ("The second branch — it repeats the first condition; the customer sees Free next-day delivery",
     "The else — the customer sees Free standard delivery",
     "The if and the elif both — everyone sees Paid delivery",
     "None — all three run and two lines print"),
    "A")

# Q30 - selecting the correct ternary at a boundary
add(2, "medium", "apply", "conditional-ternary-expressions",
    "A quiz marks a score of exactly 60 as \"Qualified\" — 60 counts as a pass. Of the lines below, one fails a score of exactly 60, one is not valid syntax, and one swaps the labels:\n\n"
    "```python\nscore = 60\nstatus = ________\nprint(status)\n```\n\n"
    "Which line is correct?",
    "Since 60 passes, the test needs `>=`, with the labels in the right order: `\"Qualified\" if score >= 60 else \"Try again\"`. Line A uses `>`, which fails a score of exactly 60; line C is statement syntax, not an expression; line D puts the labels the wrong way round.",
    ("\"Qualified\" if score > 60 else \"Try again\"",
     "\"Qualified\" if score >= 60 else \"Try again\"",
     "if score >= 60: \"Qualified\" else: \"Try again\"",
     "\"Try again\" if score >= 60 else \"Qualified\""),
    "B")

# Q31 - completing a strict less-than condition (boundary value in code)
add(2, "easy", "apply", "the-if-statement",
    "A store should alert to reorder when stock drops below 20, but stay quiet when stock is exactly 20. The stock below is exactly 20:\n\n"
    "```python\nstock = 20\nif ________:\n    print(\"Reorder now\")\n```\n\n"
    "Which condition stays quiet at 20 but would alert at 19?",
    "\"Below 20\" means strictly less than 20, so `stock < 20` is `20 < 20` → False and stays quiet, while 19 would alert. `stock <= 20` would wrongly alert at 20, `stock > 20` reverses the rule, and `stock == 20` matches only 20.",
    ("stock <= 20",
     "stock > 20",
     "stock < 20",
     "stock == 20"),
    "C")

# Q32 - equivalence: two ifs with opposite conditions vs if-else
add(2, "hard", "analyze", "if-else",
    "A reviewer worries the two-`if` heater code might set the value twice, or disagree with the `if`/`else` version for some temperature:\n\n"
    "```python\n# Version A\nif temp < 18:\n    heater = \"ON\"\nelse:\n    heater = \"OFF\"\n\n# Version B\nif temp < 18:\n    heater = \"ON\"\nif temp >= 18:\n    heater = \"OFF\"\n```\n\n"
    "The two conditions are opposites. Can the versions ever give different results?",
    "For any temperature, exactly one of `temp < 18` and `temp >= 18` is True — they are exact opposites that cover every case. So both versions set the same value every time. Version A is still safer, because an `else` cannot drift out of step during later edits.",
    ("Yes — they differ at temp == 18",
     "Yes — they differ at temp == 17",
     "Yes — two ifs always set the value twice",
     "No — one of the two opposite conditions is always true, so both set the same value"),
    "D")

# Q33 - equivalence with a subtle inner else
add(2, "hard", "analyze", "nested-conditions-and-guard-clauses",
    "These two fraud checks look the same, but notice that Version A's inner `if` has an `else`. A reviewer tries both with a small ₹5000 payment:\n\n"
    "```python\n# Version A\nif amount > 10000:\n    if foreign:\n        print(\"Flag for review\")\n    else:\n        print(\"Domestic large payment\")\n\n# Version B\nif amount > 10000 and foreign:\n    print(\"Flag for review\")\nelse:\n    print(\"Domestic large payment\")\n```\n\n"
    "For the ₹5000 payment, do they give the same result?",
    "No. For ₹5000, Version A's outer `if` is False, so nothing prints. Version B's combined test is also False, so its `else` runs and prints `Domestic large payment`. Version A's inner `else` only covers a large payment that is not foreign; Version B's `else` covers everything the combined test misses.",
    ("No — Version A prints nothing, but Version B prints Domestic large payment",
     "Yes — they always match",
     "No — Version A flags small foreign payments",
     "No — Version B never checks foreign"),
    "A")

# Q34 - final value from a match block
add(2, "easy", "understand", "match-case",
    "A parking gate reads the vehicle type \"bike\" and runs it through this match block:\n\n"
    "```python\nvehicle = \"bike\"\nmatch vehicle:\n    case \"car\":\n        fee = 50\n    case \"bike\":\n        fee = 20\n    case \"truck\":\n        fee = 100\n    case _:\n        fee = 0\nprint(fee)\n```\n\n"
    "What is fee just before it prints?",
    "`match` checks each case from top to bottom for an exact match and runs the first one that fits. `\"bike\"` matches the second case, so `fee = 20`. The `case _` default is only used when no named case matches.",
    ("0",
     "20",
     "50",
     "100"),
    "B")

# Q35 - validation with .isdigit()
add(2, "medium", "apply", "input-validation",
    "An ATM checks the typed amount before turning it into a number. The field holds \"50x\":\n\n"
    "```python\nentry = \"50x\"\nif not entry.isdigit():\n    print(\"Enter numbers only\")\nelse:\n    print(\"Processing\", int(entry))\n```\n\n"
    "What does the code do?",
    "`\"50x\".isdigit()` is False because the text is not all digits, so `not False` → True and the guard runs, printing `Enter numbers only`. The `int(entry)` on the `else` side never runs — which is exactly the crash the check prevents.",
    ("Prints Processing 50 — isdigit ignores the letter",
     "Crashes on int() with an error",
     "Prints Enter numbers only — \"50x\" is not all digits",
     "Prints Processing 50x"),
    "C")

# Q36 - tracing guard clauses in order
add(2, "medium", "understand", "input-validation",
    "An uploader checks its rules in order for an 8 MB file named \"report.pdf\". The size limit is 5 MB:\n\n"
    "```python\nname = \"report.pdf\"\nsize = 8\n\nif not name:\n    print(\"File name required\")\nelif size > 5:\n    print(\"File too large\")\nelse:\n    print(\"Upload accepted\")\n```\n\n"
    "Which check stops the upload, and what prints?",
    "The first check `not name` is False, because `\"report.pdf\"` is a non-empty (truthy) name. The `elif size > 5` is `8 > 5` → True, so it runs and prints `File too large`. The `else` runs only when both checks pass.",
    ("Nothing stops it — Upload accepted",
     "The name check stops it — File name required",
     "It crashes comparing size to 5",
     "The size check stops it — File too large"),
    "D")

# Q37 - truthiness of None
add(2, "easy", "understand", "truthiness-and-boolean-logic",
    "A new profile has no photo yet, so `photo` is `None`. A designer expects the default avatar to show:\n\n"
    "```python\nphoto = None\nif photo:\n    print(\"Showing photo\")\nelse:\n    print(\"Showing default avatar\")\n```\n\n"
    "What prints?",
    "`None` is falsy, so `if photo:` is False and the `else` runs, printing `Showing default avatar` — just as the designer expects. `None` is Python's way of saying \"no value\", so it counts as absent rather than causing an error.",
    ("Showing photo — None is truthy",
     "Showing default avatar — None is falsy",
     "An error — photo has no value",
     "Nothing — None skips both branches"),
    "B")

# Q38 - analysing a nested chain
add(2, "medium", "understand", "nested-conditions-and-guard-clauses",
    "A checked-in passenger arrives with 2 bags. The free allowance is up to 2 bags:\n\n"
    "```python\nchecked_in = True\nbags = 2\n\nif checked_in:\n    if bags == 0:\n        print(\"No baggage\")\n    elif bags <= 2:\n        print(\"Within free allowance\")\n    else:\n        print(\"Excess baggage fee\")\nelse:\n    print(\"Please check in first\")\n```\n\n"
    "Which message shows?",
    "The passenger is checked in, so the inner chain runs. `bags == 0` is False, then `bags <= 2` is `2 <= 2` → True, so it prints `Within free allowance`. The inner `else` and the outer `else` are both skipped.",
    ("Within free allowance",
     "Excess baggage fee",
     "No baggage",
     "Please check in first"),
    "A")

# Q39 - operator precedence with no brackets
add(2, "hard", "analyze", "truthiness-and-boolean-logic",
    "A notification rule has no brackets. Here `urgent` is True, `weekend` is False, and `muted` is True:\n\n"
    "```python\nweekend = False\nurgent = True\nmuted = True\nif urgent or weekend and not muted:\n    print(\"Notify user\")\nelse:\n    print(\"Stay silent\")\n```\n\n"
    "What prints? (Order: `not` first, then `and`, then `or`.)",
    "Work inside-out: `weekend and not muted` → `False and (not True)` → `False and False` → False. The line becomes `urgent or False` → `True or False` → True, so it prints `Notify user`. `or` needs only one side to be True.",
    ("Stay silent — being muted stops all alerts",
     "Nothing — the operators cancel out",
     "Notify user — urgent is True, and or needs only one true side",
     "An error — you cannot mix and and or"),
    "C")

# Q40 - truthiness in a compound and condition
add(2, "medium", "apply", "truthiness-and-boolean-logic",
    "A signup button lights only when both the name and email are filled. The name is \"Ravi\", but the email box was left empty:\n\n"
    "```python\nname = \"Ravi\"\nemail = \"\"\nif name and email:\n    print(\"Form ready\")\nelse:\n    print(\"Complete all fields\")\n```\n\n"
    "What does the check decide?",
    "`name` is a non-empty (truthy) string, but `email` is `\"\"`, which is falsy. `and` needs both sides truthy, so `\"Ravi\" and \"\"` is False overall and the `else` runs, printing `Complete all fields`. One empty field is enough to fail the check.",
    ("Form ready — the name is filled, that is enough",
     "An error — you cannot use and with strings",
     "Form ready — a filled name makes the whole thing true",
     "Complete all fields — the empty email is falsy, so the and is False"),
    "D")


# =====================  Build workbook  =====================

assert len(Q) == 40, f"expected 40 questions, got {len(Q)}"

os.makedirs(OUT_DIR, exist_ok=True)
shutil.copyfile(TEMPLATE, OUT_FILE)

wb = openpyxl.load_workbook(OUT_FILE)
ws = wb.active
ws.title = SHEET_NAME
ws.delete_rows(2, ws.max_row)  # keep header, drop sample rows

set_counter = {1: 0, 2: 0}
letters = []
answer_tally = Counter()

for setn, diff, bloom, sub, desc, expl, opts, ans in Q:
    set_counter[setn] += 1
    title = f"Python - MCQ - 3.{setn}.{set_counter[setn]}"
    tag = f"python - Set {setn}"
    answer_tally[ans] += 1
    letters.append(ans)
    ws.append([
        title, desc, expl, SCORE[diff], "published", diff, bloom,
        tag, "python", "control-flow", sub, None,
        opts[0], opts[1], opts[2], opts[3], L[ans],
    ])

wb.save(OUT_FILE)

# =====================  Validation  =====================
print("Saved:", OUT_FILE)
print("Total questions:", len(Q))
print("Answer distribution:", dict(sorted(answer_tally.items())))

run = max_run = 1
for a, b in zip(letters, letters[1:]):
    run = run + 1 if a == b else 1
    max_run = max(max_run, run)
print("Max consecutive same letter:", max_run)

assert all(answer_tally[x] == 10 for x in "ABCD"), answer_tally
assert max_run <= 2, "answer letter repeats more than twice consecutively"

chk = openpyxl.load_workbook(OUT_FILE)
cs = chk[SHEET_NAME]
rows = list(cs.iter_rows(values_only=True))
assert rows[0] == ('title', 'description', 'explanation', 'score', 'status',
                   'difficulty', 'bloomTaxonomy', 'tags', 'subjects', 'topics',
                   'subTopics', 'companies', 'option1', 'option2', 'option3',
                   'option4', 'answer'), rows[0]
assert len(rows) == 41, len(rows)
for r in rows[1:]:
    assert len(set(r[12:16])) == 4, f"duplicate options in {r[0]}"
    assert r[12:16][r[16] - 1], f"empty correct option in {r[0]}"
print("Difficulty mix:", dict(Counter(r[5] for r in rows[1:])))
print("Set 1:", sum(1 for r in rows[1:] if r[7] == 'python - Set 1'),
      "| Set 2:", sum(1 for r in rows[1:] if r[7] == 'python - Set 2'))
print("Validation passed.")
