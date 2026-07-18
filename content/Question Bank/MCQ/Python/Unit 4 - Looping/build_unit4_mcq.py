"""Build the Unit 4 - Looping MCQ workbook from the shared template.

Grounded strictly in content/Semester 1/Unit 4 - Looping:
why loops (for = known count / every item, while = unknown count until a
condition changes) / while loops (counter pattern, sentinel pattern, forgetting
the update -> infinite loop) / for loops and range (three forms; stop is
excluded) / iterating over strings and lists, enumerate(..., start=1) /
break, continue, and the loop-else clause / nested loops (inner runs fully per
outer pass; total = outer * inner) / the four common patterns (count, sum/
average with len(), min/max best-so-far, search with a flag + break; while True
+ break sentinel; guard against divide-by-zero) / avoiding infinite loops and
off-by-one errors (range excludes stop, < vs <=, trace first and last pass).

Writing rules (same bar as Units 1 and 3):
* Simple English, short sentences.
* Trace-values sit on or near boundaries; a wrong reasoning path gives a
  different answer, so the student must actually work it out.
* Each distractor maps to one specific, nameable misconception.
* Every question demands its own distinct reasoning step.
* Only loop features taught in this unit are used (f-strings, print(end=...),
  range, enumerate, len, break/continue/else, %, lists/strings to iterate).

Set 1 (Q1-10)  -> tag "python - Set 1", titles 4.1.1 .. 4.1.10
Set 2 (Q11-40) -> tag "python - Set 2", titles 4.2.1 .. 4.2.30
Answer field is 1..4 (A=1, B=2, C=3, D=4).
"""

import os
import shutil
from collections import Counter

import openpyxl

TEMPLATE = "content/Question Bank/Template/questions-mcq-template.xlsx"
OUT_DIR = "content/Question Bank/MCQ/Python/Unit 4 - Looping"
OUT_FILE = os.path.join(OUT_DIR, "Unit 4 - Looping - MCQ.xlsx")
SHEET_NAME = "Python - MCQ - 4"

L = {"A": 1, "B": 2, "C": 3, "D": 4}
SCORE = {"easy": 5, "medium": 8, "hard": 10}

Q = []


def add(setn, diff, bloom, sub, desc, expl, opts, ans):
    Q.append((setn, diff, bloom, sub, desc, expl, opts, ans))


# =====================  SET 1  (Q1-Q10)  =====================

# Q1 - range excludes its stop value (off-by-one surprise)
add(1, "medium", "analyze", "for-loops-and-range",
    "A student wants to print the numbers 1 to 5 and writes this, but the result surprises them:\n\n"
    "```python\nfor i in range(1, 5):\n    print(i)\n```\n\n"
    "What does it print, and why?",
    "`range(1, 5)` starts at 1 and stops just before 5, so it never includes the stop value. It prints 1, 2, 3, 4. To include 5 the student would need `range(1, 6)`.",
    ("1 2 3 4 5 — range includes both ends",
     "0 1 2 3 4 — range always starts at 0",
     "1 2 3 4 — range stops just before 5, so 5 is left out",
     "2 3 4 5 — range skips the start value"),
    "C")

# Q2 - while loop with no counter update -> infinite (lesson 08 headline example)
add(1, "medium", "analyze", "avoiding-infinite-loops-and-off-by-one-errors",
    "This loop is meant to print 1, 2, 3:\n\n"
    "```python\ncount = 1\nwhile count <= 3:\n    print(count)\n```\n\n"
    "What actually happens when it runs?",
    "Nothing inside the loop changes `count`, so it stays 1 forever. The condition `count <= 3` never becomes false, so the loop prints 1 endlessly — an infinite loop. Adding `count = count + 1` inside the loop fixes it.",
    ("It prints 1 over and over forever, because count never changes",
     "It prints 1 2 3 then stops",
     "It prints nothing, because count is never updated",
     "It prints 1 once and stops"),
    "A")

# Q3 - choosing while for an unknown count
add(1, "medium", "understand", "why-loops-repetition-and-automation",
    "Kabir wants to keep asking for a wifi password until the user finally types the correct one. He has no idea whether it will take two tries or twelve.\n\n"
    "Which loop suits this best, and why?",
    "When the number of repeats is unknown and you just keep going until a condition changes, a `while` loop fits. A `for` loop is for a known count or visiting every item in a sequence, which is not the case here.",
    ("A for loop, because the number of tries is known",
     "A for loop with range(), because it is always the safer choice",
     "Either works the same; there is no real difference",
     "A while loop, because the count is unknown and it repeats until the password is correct"),
    "D")

# Q4 - looping over a string
add(1, "easy", "understand", "iterating-over-sequences-and-strings",
    "What does this loop print?\n\n"
    "```python\nfor letter in \"cat\":\n    print(letter)\n```",
    "A string is a sequence of characters, so the `for` loop visits each one in turn. The loop variable `letter` holds a single character per pass, printing `c`, then `a`, then `t`, each on its own line.",
    ("cat, all on one line",
     "c, a, and t, each on its own line",
     "0, 1, 2 — the positions of the letters",
     "An error — you cannot loop over a string"),
    "B")

# Q5 - break stops before printing the matched value
add(1, "medium", "analyze", "break-continue-and-loop-else",
    "What does this print?\n\n"
    "```python\nfor n in range(1, 10):\n    if n == 3:\n        break\n    print(n)\n```",
    "The loop prints 1, then 2. When `n` reaches 3, the `if` is true and `break` ends the loop immediately — before the `print` on that pass runs. So 3 is never printed, and the output is 1 and 2.",
    ("1 2 3",
     "3",
     "1 2 — the loop stops when n reaches 3, before printing it",
     "1 2 3 4 5 6 7 8 9"),
    "C")

# Q6 - nested loop total = outer * inner
add(1, "medium", "apply", "nested-loops-grids-and-patterns",
    "An outer loop runs 4 times, and for each pass of the outer loop an inner loop runs 5 times.\n\n"
    "How many times does the inner block run in total?",
    "The inner loop runs fully for every pass of the outer loop, so the total is the outer count times the inner count: 4 × 5 = 20. Adding them (9) or taking just one count misses how the repeats multiply.",
    ("20 — the outer count times the inner count",
     "9 — the outer count plus the inner count",
     "5 — only the inner count matters",
     "4 — only the outer count matters"),
    "A")

# Q7 - sentinel pattern
add(1, "easy", "understand", "while-loops-counters-and-sentinels",
    "Look at this loop:\n\n"
    "```python\ncommand = \"\"\nwhile command != \"quit\":\n    command = input(\"Command: \")\n```\n\n"
    "What makes this loop stop?",
    "This is the sentinel pattern: the loop keeps reading commands and only stops when a special signal value appears. Here that value is `\"quit\"` — once the user types it, the condition becomes false and the loop ends.",
    ("It stops after a fixed number of commands",
     "It never stops — while loops always run forever",
     "It stops as soon as the user types anything",
     "It stops when the user types \"quit\", the sentinel value"),
    "D")

# Q8 - continue skips the current pass
add(1, "hard", "analyze", "break-continue-and-loop-else",
    "What does this print?\n\n"
    "```python\nfor n in range(1, 6):\n    if n % 2 == 0:\n        continue\n    print(n)\n```",
    "`range(1, 6)` gives 1, 2, 3, 4, 5. When `n` is even, `n % 2 == 0` is true and `continue` skips the rest of that pass, so the even numbers are never printed. Only the odd numbers 1, 3, 5 appear.",
    ("1 2 3 4 5",
     "1 3 5 — continue skips the even numbers",
     "2 4 — only the even numbers print",
     "Nothing — continue skips every number"),
    "B")

# Q9 - min/max pattern trace
add(1, "hard", "analyze", "common-loop-patterns",
    "This loop finds the largest value:\n\n"
    "```python\nnumbers = [42, 17, 88, 23]\nlargest = numbers[0]\nfor n in numbers:\n    if n > largest:\n        largest = n\nprint(largest)\n```\n\n"
    "What is printed, and how does it work?",
    "It uses the best-so-far pattern: start by assuming the first item (42) is the largest, then replace `largest` whenever a bigger value appears. 88 is bigger, so it takes over; 23 does not. The result is 88.",
    ("88 — it assumes the first is largest, then replaces it whenever a bigger value appears",
     "42 — largest keeps its first value the whole time",
     "23 — the last value always wins",
     "170 — it adds all the values together"),
    "A")

# Q10 - range with a step
add(1, "medium", "analyze", "for-loops-and-range",
    "What does this print?\n\n"
    "```python\nfor i in range(0, 10, 2):\n    print(i)\n```",
    "`range(0, 10, 2)` starts at 0 and counts up in steps of 2, stopping just before 10. So it produces 0, 2, 4, 6, 8. The stop value 10 is excluded, just like every range.",
    ("0 1 2 3 4 5 6 7 8 9",
     "2 4 6 8 10",
     "0 2 4 6 8 — it counts in steps of 2 and stops before 10",
     "0 2 4 6 8 10 — it includes the stop value"),
    "C")


# =====================  SET 2  (Q11-Q40)  =====================

# Q11 - while with < stops one before
add(2, "medium", "analyze", "avoiding-infinite-loops-and-off-by-one-errors",
    "Which numbers does this print?\n\n"
    "```python\ni = 1\nwhile i < 5:\n    print(i)\n    i = i + 1\n```",
    "The loop runs while `i` is strictly less than 5. It prints 1, 2, 3, 4; when `i` becomes 5, `5 < 5` is false and the loop stops. Using `<=` instead would also print 5.",
    ("1 2 3 4 5 — it includes 5",
     "1 2 3 4 — it stops before 5 because the test uses <",
     "1 2 3 — it stops one number early",
     "It runs forever"),
    "B")

# Q12 - enumerate with start=1
add(2, "medium", "analyze", "iterating-over-sequences-and-strings",
    "Look at this loop:\n\n"
    "```python\nnames = [\"Asha\", \"Ravi\", \"Meera\"]\nfor position, name in enumerate(names, start=1):\n    print(position, name)\n```\n\n"
    "What is the first line it prints?",
    "`enumerate` hands you the position and the item together, and `start=1` makes the numbering begin at 1 instead of 0. The first item is `\"Asha\"`, paired with position 1, so the first line is `1 Asha`.",
    ("0 Asha — enumerate always starts at 0",
     "Asha 1 — the name comes before the number",
     "1 Meera — enumerate starts from the last item",
     "1 Asha — start=1 numbers from 1, paired with the first item"),
    "D")

# Q13 - counting pattern
add(2, "medium", "analyze", "common-loop-patterns",
    "What is printed?\n\n"
    "```python\nnumbers = [4, 9, 2, 7, 6, 1]\nevens = 0\nfor n in numbers:\n    if n % 2 == 0:\n        evens = evens + 1\nprint(evens)\n```",
    "This is the counting pattern: `evens` starts at 0 and rises by 1 only when a number is even. The even numbers are 4, 2, and 6, so the counter ends at 3.",
    ("3 — it counts how many numbers are even",
     "6 — it counts all the numbers",
     "12 — it adds the even numbers together",
     "0 — the counter never changes"),
    "A")

# Q14 - loop-else runs when no break
add(2, "hard", "analyze", "break-continue-and-loop-else",
    "What prints, and why?\n\n"
    "```python\nfor n in [2, 4, 6, 8]:\n    if n == 7:\n        print(\"Found it!\")\n        break\nelse:\n    print(\"Not found.\")\n```",
    "A loop's `else` runs only if the loop finished without ever hitting a `break`. 7 is not in the list, so the `break` never fires and the loop ends normally — so the `else` runs and prints `Not found.`",
    ("Found it! — 7 matches one of the numbers",
     "Nothing — the else never runs after a for loop",
     "Not found. — the loop never breaks, so the else runs",
     "Both Found it! and Not found."),
    "C")

# Q15 - sum / average with len()
add(2, "medium", "analyze", "common-loop-patterns",
    "What does this print?\n\n"
    "```python\nnumbers = [10, 20, 30, 40]\ntotal = 0\nfor n in numbers:\n    total = total + n\nprint(total / len(numbers))\n```",
    "`total` adds up to 10 + 20 + 30 + 40 = 100. `len(numbers)` is 4, so the average printed is 100 / 4 = 25.0. `len()` counts the items so you do not have to.",
    ("100 — it prints the total, not the average",
     "25.0 — the sum is 100, divided by the 4 items",
     "4 — len just counts the items",
     "40 — only the last value is kept"),
    "B")

# Q16 - while True + break is deliberate, not a bug
add(2, "medium", "understand", "avoiding-infinite-loops-and-off-by-one-errors",
    "A programmer writes `while True:` on purpose and puts a `break` inside to leave the loop at the right moment.\n\n"
    "Is this a bug?",
    "`while True:` with a `break` is a deliberate, controlled loop — you loop on purpose and escape when the right moment comes (for example, when the user types \"done\"). It is only an accidental infinite loop when there is no way out.",
    ("Yes — while True always freezes the program",
     "Yes — you can never leave a while True loop",
     "No, but only because Python stops it automatically after a while",
     "No — it is a deliberate, controlled loop that break exits when needed"),
    "D")

# Q17 - nested star triangle
add(2, "hard", "analyze", "nested-loops-grids-and-patterns",
    "What shape does this print?\n\n"
    "```python\nfor row in range(1, 4):\n    for star in range(row):\n        print(\"*\", end=\"\")\n    print()\n```",
    "For each outer pass, the inner loop runs `row` times. Row 1 prints one star, row 2 prints two, row 3 prints three. `end=\"\"` keeps stars on the same line, and the empty `print()` starts a new line — giving a growing triangle.",
    ("A growing triangle: one star, then two, then three, each on its own line",
     "A solid 3 by 3 square of stars",
     "Three stars on a single line: ***",
     "Nothing — you cannot nest loops"),
    "A")

# Q18 - for is safer for counted jobs
add(2, "medium", "understand", "for-loops-and-range",
    "When either loop could work for a counted job, the unit recommends a for loop over a hand-managed while loop.\n\n"
    "What is the main reason?",
    "A `for` loop does the counting itself — there is no counter to set up and no update to remember — so it cannot accidentally become infinite. A hand-managed `while` can loop forever if you forget to update its variable.",
    ("A for loop always runs faster than a while loop",
     "A while loop cannot use range()",
     "A for loop handles the counting itself, so it cannot accidentally become infinite",
     "A while loop cannot contain an if statement"),
    "C")

# Q19 - search pattern with a flag
add(2, "medium", "analyze", "common-loop-patterns",
    "What is printed?\n\n"
    "```python\nnumbers = [3, 8, 5, 9]\ntarget = 5\nfound = False\nfor n in numbers:\n    if n == target:\n        found = True\n        break\nprint(found)\n```",
    "This is the search pattern. The flag `found` starts False and flips to True only if the target appears. 5 is in the list, so `found` becomes True and `break` stops the search. The output is `True`.",
    ("False — the flag never changes",
     "True — the target 5 is in the list, so found flips to True",
     "5 — it prints the target value",
     "An error — you cannot print a flag"),
    "B")

# Q20 - range(stop) basic
add(2, "easy", "understand", "for-loops-and-range",
    "What does this print?\n\n"
    "```python\nfor i in range(5):\n    print(i)\n```",
    "`range(5)` produces 0, 1, 2, 3, 4 — it starts at 0 and stops just before 5. So the loop prints those five numbers.",
    ("1 2 3 4 5",
     "0 1 2 3 4 5",
     "5",
     "0 1 2 3 4 — range(5) goes from 0 up to but not including 5"),
    "D")

# Q21 - the most-forgotten part of the counter pattern
add(2, "medium", "understand", "while-loops-counters-and-sentinels",
    "The counter pattern for a while loop has three parts: set up the counter, test it in the condition, and one more. Leaving out that last part causes an infinite loop.\n\n"
    "Which part is most often forgotten?",
    "The three parts are setup, test, and update. Forgetting to update the counter inside the loop means the condition never changes, so the loop runs forever. The other choices are not required parts of the pattern.",
    ("Updating the counter inside the loop",
     "Printing the counter each pass",
     "Writing a comment above the loop",
     "Choosing a good name for the counter"),
    "A")

# Q22 - nested loop growth on large data
add(2, "medium", "understand", "nested-loops-grids-and-patterns",
    "An outer loop runs 1,000 times, and its inner loop runs 1,000 times for each outer pass.\n\n"
    "Roughly how many times does the inner block run, and why is that worth watching?",
    "The inner block runs outer × inner = 1,000 × 1,000 = 1,000,000 times. This multiplication is fine for small data but grows very fast, which is why nested loops need care on large amounts of data.",
    ("2,000 — you add the two counts",
     "1,000 — only the outer loop counts",
     "1,000,000 — outer times inner, which grows fast for large data",
     "1,000,000,000 — outer raised to the power of inner"),
    "C")

# Q23 - continue vs break
add(2, "medium", "understand", "break-continue-and-loop-else",
    "Inside a loop you want to skip the current item but keep going with the rest.\n\n"
    "Which keyword does that, and how is it different from the other?",
    "`continue` abandons only the current pass and jumps to the next item, while the loop keeps running. `break` is different — it stops the whole loop immediately. So `continue` skips, `break` stops.",
    ("break — it skips one item and then continues",
     "continue — it skips only the current pass; break would stop the whole loop",
     "loop else — it skips the current item",
     "return — it skips the current item"),
    "B")

# Q24 - vowel count trace (loop + if)
add(2, "hard", "analyze", "iterating-over-sequences-and-strings",
    "What is printed?\n\n"
    "```python\nword = \"melon\"\nvowels = 0\nfor letter in word:\n    if letter in \"aeiou\":\n        vowels = vowels + 1\nprint(vowels)\n```",
    "The loop checks each letter of `melon`. `m` is not a vowel, `e` is, `l` is not, `o` is, `n` is not. So `vowels` counts e and o, ending at 2.",
    ("5 — it counts every letter",
     "3 — m, l, and n are counted",
     "0 — the counter never changes",
     "2 — only e and o are vowels"),
    "D")

# Q25 - smallest fix for range off-by-one
add(2, "medium", "apply", "avoiding-infinite-loops-and-off-by-one-errors",
    "A loop `for i in range(1, 5)` prints 1 to 4, but you actually wanted 1 to 5 included.\n\n"
    "What is the smallest correct fix?",
    "Because `range` excludes its stop value, `range(1, 5)` stops at 4. To include 5, raise the stop to 6: `range(1, 6)`. Changing the start or printing `i + 1` would shift the numbers wrongly.",
    ("Change range(1, 5) to range(1, 6)",
     "Change range(1, 5) to range(0, 5)",
     "Add a second loop after this one",
     "Change print(i) to print(i + 1)"),
    "A")

# Q26 - while condition false at the start -> body never runs
add(2, "hard", "analyze", "while-loops-counters-and-sentinels",
    "What does this print?\n\n"
    "```python\ncount = 10\nwhile count < 5:\n    print(count)\n    count = count + 1\nprint(\"done\")\n```",
    "A `while` loop checks its condition before each pass. Here `10 < 5` is already false, so the loop body never runs at all. Only the line after the loop runs, printing `done`.",
    ("10 done",
     "It runs forever",
     "done — the condition is false at the start, so the body never runs",
     "5 6 7 8 9 done"),
    "C")

# Q27 - choosing the counting pattern
add(2, "medium", "apply", "common-loop-patterns",
    "Kabir wants to know how many classmates on a list scored above 80. He plans to look at each score and keep a running tally.\n\n"
    "Which loop pattern is this?",
    "Keeping a tally that starts at 0 and rises by 1 each time a condition is true is the counting pattern. The sum pattern would add the scores, min/max would track a best value, and search would just check whether one exists.",
    ("The sum / average pattern",
     "The counting pattern — start at 0 and add 1 each time the condition is true",
     "The min / max pattern",
     "The search pattern"),
    "B")

# Q28 - min/max wrong starting value
add(2, "hard", "analyze", "common-loop-patterns",
    "A loop that finds the largest score starts with `largest = 0`. It works for positive scores, but a colleague warns it can give a wrong answer.\n\n"
    "When would starting at 0 be wrong?",
    "The safe start is the first item in the data, not 0. If every value is below 0, no value ever beats 0, so `largest` wrongly stays at 0 — a value that was never even in the list. Starting from `numbers[0]` avoids this.",
    ("Never — starting at 0 is always safe",
     "When the list is very long",
     "When the values are all above 0",
     "When every value is below 0 — largest would wrongly stay at 0"),
    "D")

# Q29 - sentinel gather + divide-by-zero guard
add(2, "hard", "analyze", "common-loop-patterns",
    "The user types \"done\" straight away, before entering any numbers:\n\n"
    "```python\ntotal = 0\ncount = 0\nwhile True:\n    entry = input(\"Enter a number (or 'done'): \")\n    if entry == \"done\":\n        break\n    total = total + int(entry)\n    count = count + 1\nif count > 0:\n    print(total / count)\nelse:\n    print(\"No numbers entered.\")\n```\n\n"
    "What prints, and why is the count > 0 check there?",
    "Typing \"done\" first breaks the loop immediately, so `count` stays 0. The `if count > 0` guard then sends the program to the `else`, printing `No numbers entered.` The guard is there to avoid dividing by zero, which would crash.",
    ("No numbers entered. — count is 0, and the check avoids dividing by zero",
     "0 — the average of no numbers is 0",
     "An error — dividing by zero crashes the program",
     "done — it prints the sentinel value"),
    "A")

# Q30 - enumerate gives index + item
add(2, "medium", "understand", "iterating-over-sequences-and-strings",
    "You are looping over a list of names and want both the name and its position number, without going back to range() and index numbers.\n\n"
    "Which tool gives you both at once?",
    "`enumerate` hands you the position and the item together on each pass, so you never manage index numbers yourself. `len` only counts items, and `break`/`continue` control the flow rather than giving positions.",
    ("len()",
     "break",
     "enumerate()",
     "continue"),
    "C")

# Q31 - for over a list, one pass per item
add(2, "easy", "understand", "iterating-over-sequences-and-strings",
    "How many lines does this print?\n\n"
    "```python\nnames = [\"Asha\", \"Ravi\", \"Meera\"]\nfor name in names:\n    print(\"Hello,\", name)\n```",
    "A `for` loop runs its body once for each item in the list. There are three names, so it greets each one — three lines. It does not print the whole list at once, nor one line per letter.",
    ("1 — it prints the whole list at once",
     "0 — you cannot loop over a list",
     "9 — one line for each letter",
     "3 — one line for each name in the list"),
    "D")

# Q32 - diagnosing / stopping an infinite loop
add(2, "medium", "understand", "avoiding-infinite-loops-and-off-by-one-errors",
    "A while loop seems to freeze the program and never stop.\n\n"
    "What is the most likely cause, and how do you stop a runaway program?",
    "A frozen program usually means the `while` condition never becomes false, so the loop never ends. You can stop a runaway program by pressing Ctrl + C in the terminal (or the stop button in your editor), then fix the loop so each pass makes progress.",
    ("The loop has too many print statements; delete some",
     "Its condition never becomes false; press Ctrl + C to stop it",
     "range() was used incorrectly; switch to a for loop",
     "The loop body is indented; remove the indentation"),
    "B")

# Q33 - trace first/last pass, sum 1..9
add(2, "hard", "apply", "avoiding-infinite-loops-and-off-by-one-errors",
    "Trace this by hand before deciding:\n\n"
    "```python\ntotal = 0\ni = 1\nwhile i < 10:\n    total = total + i\n    i = i + 1\nprint(total)\n```\n\n"
    "What does total add up to?",
    "On the first pass `i` is 1; the loop keeps going while `i < 10`, so the last value added is 9 (when `i` becomes 10 the loop stops). It adds 1 through 9, which is 45 — not 1 through 10.",
    ("45 — it adds 1 through 9, because the loop stops before i reaches 10",
     "55 — it adds 1 through 10",
     "10 — it prints the final value of i",
     "It runs forever"),
    "A")

# Q34 - break to end a search
add(2, "medium", "apply", "break-continue-and-loop-else",
    "You are searching a list and want to stop the moment you find a match, with no reason to keep looking.\n\n"
    "Which keyword fits?",
    "`break` stops the whole loop immediately, which is exactly what you want once a search has succeeded. `continue` would only skip one item, and the loop `else` runs when there was no break at all.",
    ("continue — it ends the search",
     "loop else — it ends the search",
     "break — it stops the whole loop once the match is found",
     "enumerate — it stops the search"),
    "C")

# Q35 - nested loop grid layout
add(2, "medium", "analyze", "nested-loops-grids-and-patterns",
    "How many numbers does this print, and how are they laid out?\n\n"
    "```python\nfor a in range(1, 4):\n    for b in range(1, 4):\n        print(a * b, end=\" \")\n    print()\n```",
    "The outer loop runs 3 times (a = 1, 2, 3) and the inner loop runs 3 times each (b = 1, 2, 3), so 3 × 3 = 9 numbers print. `end=\" \"` keeps a row on one line, and the empty `print()` ends each row — a 3 by 3 grid.",
    ("3 numbers on one line",
     "6 numbers in two rows",
     "12 numbers in a 3 by 4 grid",
     "9 numbers in a 3 by 3 grid, one row per outer pass"),
    "D")

# Q36 - range(start, stop) form
add(2, "easy", "understand", "for-loops-and-range",
    "What sequence of numbers does `range(2, 7)` produce?",
    "`range(2, 7)` starts at 2 and stops just before 7, so it produces 2, 3, 4, 5, 6. The stop value 7 is excluded, as in every range.",
    ("2 3 4 5 6 7",
     "2 3 4 5 6 — from 2 up to but not including 7",
     "0 1 2 3 4 5 6",
     "2 7 — just the start and stop"),
    "B")

# Q37 - for is best when the count is known
add(2, "medium", "apply", "why-loops-repetition-and-automation",
    "Which of these tasks is best done with a for loop rather than a while loop?",
    "A `for` loop suits a known count or visiting every item. Printing 1 to 50 has a known count, so `for` fits. The other tasks repeat until something changes with no fixed count, which is `while` territory.",
    ("Printing every number from 1 to 50 — the count is known",
     "Rolling a dice until it lands on 6 — the count is unknown",
     "Asking for a password until it is correct — the count is unknown",
     "Keeping a game running until the player quits — the count is unknown"),
    "A")

# Q38 - loop-else skipped when break fires
add(2, "medium", "analyze", "break-continue-and-loop-else",
    "What prints here?\n\n"
    "```python\nfor n in [2, 4, 5, 8]:\n    if n == 5:\n        print(\"Found\")\n        break\nelse:\n    print(\"Not found\")\n```",
    "5 is in the list, so on that pass the `if` is true: it prints `Found` and `break`s out. Because the loop ended with a `break`, the `else` is skipped. Only `Found` prints.",
    ("Found and Not found",
     "Not found",
     "Found — the break fires, so the else is skipped",
     "Nothing"),
    "C")

# Q39 - < vs <= differ by one
add(2, "medium", "analyze", "avoiding-infinite-loops-and-off-by-one-errors",
    "Two loops are identical except one uses `while i < 5` and the other `while i <= 5`. Both start at `i = 1` and add 1 each pass.\n\n"
    "How do their results differ?",
    "`< 5` stops once `i` reaches 5, so the last pass uses i = 4. `<= 5` allows one more pass with i = 5 before stopping. So the `<=` version runs exactly one extra time — the classic off-by-one difference.",
    ("They behave exactly the same",
     "The < version runs one time more",
     "The <= version stops one step earlier",
     "The <= version runs one extra pass, including i = 5"),
    "D")

# Q40 - why loops: write once, repeat many
add(2, "medium", "understand", "why-loops-repetition-and-automation",
    "Kabir must send the same reminder to 60 classmates. Doing it by hand is slow and easy to get wrong.\n\n"
    "What does using a loop change?",
    "A loop lets you write the repeated action once and have it run as many times as needed — here, once per classmate — quickly and without the copy-paste mistakes of doing it by hand. It does not change the message or remove the need for instructions.",
    ("It makes each message different automatically",
     "It lets him write the sending step once and repeat it for all 60, quickly and without miscopying",
     "It removes the need to know who the classmates are",
     "It sends the messages without any instructions at all"),
    "B")


# =====================  Build workbook  =====================

assert len(Q) == 40, f"expected 40 questions, got {len(Q)}"

os.makedirs(OUT_DIR, exist_ok=True)
shutil.copyfile(TEMPLATE, OUT_FILE)

wb = openpyxl.load_workbook(OUT_FILE)
ws = wb.active
ws.title = SHEET_NAME
ws.delete_rows(2, ws.max_row)  # keep header, drop sample rows

set_counter = {1: 0, 2: 0}
letters = []
answer_tally = Counter()

for setn, diff, bloom, sub, desc, expl, opts, ans in Q:
    set_counter[setn] += 1
    title = f"Python - MCQ - 4.{setn}.{set_counter[setn]}"
    tag = f"python - Set {setn}"
    answer_tally[ans] += 1
    letters.append(ans)
    ws.append([
        title, desc, expl, SCORE[diff], "published", diff, bloom,
        tag, "python", "looping", sub, None,
        opts[0], opts[1], opts[2], opts[3], L[ans],
    ])

wb.save(OUT_FILE)

# =====================  Validation  =====================
print("Saved:", OUT_FILE)
print("Total questions:", len(Q))
print("Answer distribution:", dict(sorted(answer_tally.items())))

run = max_run = 1
for a, b in zip(letters, letters[1:]):
    run = run + 1 if a == b else 1
    max_run = max(max_run, run)
print("Max consecutive same letter:", max_run)

assert all(answer_tally[x] == 10 for x in "ABCD"), answer_tally
assert max_run <= 2, "answer letter repeats more than twice consecutively"

chk = openpyxl.load_workbook(OUT_FILE)
cs = chk[SHEET_NAME]
rows = list(cs.iter_rows(values_only=True))
assert rows[0] == ('title', 'description', 'explanation', 'score', 'status',
                   'difficulty', 'bloomTaxonomy', 'tags', 'subjects', 'topics',
                   'subTopics', 'companies', 'option1', 'option2', 'option3',
                   'option4', 'answer'), rows[0]
assert len(rows) == 41, len(rows)
for r in rows[1:]:
    assert len(set(r[12:16])) == 4, f"duplicate options in {r[0]}"
    assert r[12:16][r[16] - 1], f"empty correct option in {r[0]}"
print("Difficulty mix:", dict(Counter(r[5] for r in rows[1:])))
print("Set 1:", sum(1 for r in rows[1:] if r[7] == 'python - Set 1'),
      "| Set 2:", sum(1 for r in rows[1:] if r[7] == 'python - Set 2'))
print("Validation passed.")
